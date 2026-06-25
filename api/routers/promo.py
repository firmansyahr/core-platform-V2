"""Promo program management — CRUD + monitoring endpoints."""
from __future__ import annotations

import io
import json
import os
import threading
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
import pandas as pd
from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from api.core.data_loader import load_data
from api.core.aegis_engine import get_store_crs
from api.core.promo_engine import (
    PRICE_PER_TON_ESTIMATE,
    calculate_promo_achievement,
    estimate_budget,
    get_cluster_comparison,
    get_daily_trend,
    get_promo_summary,
)
from api.core import promo_calculator as pc
from api.database import SessionLocal
from api.models import LoyaltyMember as LoyaltyMemberRow
from api.models import Promo as PromoRow
from api.models import PromoArchive as PromoArchiveRow
from api.models import PromoPeserta as PromoPesertaRow

router = APIRouter(prefix="/api/promo", tags=["promo"])

# Feature flag — Tahap 4b rollout, pola identik dengan loyalty.py (Tahap 4a).
# Default false = perilaku identik dengan sebelumnya (JSON). Lihat fungsi
# _get_promos()/_save_new_promo()/dst. di bawah untuk abstraksinya.
USE_SQLITE = os.getenv("USE_SQLITE_STORAGE", "false").lower() == "true"

# jenis_promo yang HANYA dihasilkan oleh skema v1 (_infer_jenis(), endpoint
# /create) — tidak pernah overlap dengan jenis_promo v2/v3 (multi_tier_points/
# flat_multiplier/leaderboard). Dipakai untuk merekonstruksi nama field asli
# ("konfigurasi_promo" vs "reward_config") saat baca balik dari kolom
# reward_config yang menyatukan keduanya di model Promo — lihat docstring
# class Promo di api/models.py.
_V1_JENIS_PROMO = {"reward_rate", "target_bonus", "cashback", "kombinasi"}

_LOCK = threading.Lock()


def _get_data_dir() -> Path:
    vol = Path("/mnt/data")
    if vol.exists() and os.access(vol, os.W_OK):
        d = vol / "app_data"
        d.mkdir(parents=True, exist_ok=True)
        return d
    d = Path(__file__).parent.parent / "data"
    d.mkdir(parents=True, exist_ok=True)
    return d


_DATA_DIR     = _get_data_dir()
_PROMOS_PATH  = _DATA_DIR / "promos.json"
_MEMBERS_PATH = _DATA_DIR / "loyalty_members.json"


# ── File I/O ──────────────────────────────────────────────────────────────────

def _rp(path: Path) -> Any:
    if not path.exists():
        return []
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []


def _wp(path: Path, data: Any) -> None:
    _DATA_DIR.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _meta(**kw: Any) -> dict:
    return {"generated_at": _now(), **kw}


# ── Storage abstraction (JSON / SQLite) ────────────────────────────────────────
#
# Pola identik dengan loyalty.py: setiap fungsi punya DUA cabang (USE_SQLITE
# True/False) yang mengembalikan/menerima bentuk dict YANG SAMA PERSIS dengan
# JSON asli. Cabang JSON adalah logic lama, TIDAK diubah — hanya dipindah ke
# dalam fungsi.
#
# Promo Selesai/Dibatalkan TETAP di tabel Promo yang sama (siklus hidup live
# tidak pernah "pindah tabel" — beda dari kategorisasi satu-kali di skrip
# migrasi). Tabel PromoArchive hanya berisi 1 entri legacy historis dari
# migrasi Tahap 3 (promo lama yang skemanya terlalu tidak konsisten untuk
# dinormalisasi) — read-only dari sisi aplikasi live, digabung saat baca
# supaya tetap muncul di GET /api/promo dan GET /api/promo/{id}.

def _peserta_to_dict(pp: PromoPesertaRow) -> dict:
    """rate_override/brand_utama: union shape, lihat docstring class PromoPeserta
    di api/models.py — add-one/upload-excel TIDAK PUNYA key brand_utama sama
    sekali (bukan None), monitoring-add TIDAK PUNYA key rate_override sama
    sekali. Beberapa konsumen (promo_calculator.calculate_flat_multiplier_program)
    memakai peserta.get("brand_utama", "Semen Elang") — kalau key selalu di-set
    ke None di sini, default itu TIDAK PERNAH terpakai dan reward salah hitung
    (dikonfirmasi via parity test: multiplier 2.0X jadi 1.0X). rate_override
    aman selalu disertakan karena semua pemanggil pakai .get() tanpa default
    truthy (lihat promo_engine.py)."""
    d: dict = {
        "id_toko":       pp.id_toko,
        "nama_toko":     pp.nama_toko,
        "cluster":       pp.cluster,
        "rate_override": pp.rate_override,
        "target_ton":    pp.target_ton,
        "catatan":       pp.catatan or "",
    }
    if pp.brand_utama is not None:
        d["brand_utama"] = pp.brand_utama
    return d


def _iso_utc(dt: datetime | None) -> str | None:
    """SQLite/SQLAlchemy melepas tzinfo saat baca balik kolom DateTime.
    Semua datetime yang pernah ditulis ke kolom ini berasal dari _now()
    (datetime.now(timezone.utc)) atau func.now() SQLite (juga UTC) — aman
    menempelkan kembali UTC di sini supaya output ISO string sama persis
    dengan mode JSON (yang menyimpan _now() apa adanya, termasuk +00:00)."""
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.isoformat()


def _promo_row_to_dict(p: PromoRow, peserta_rows: list[PromoPesertaRow]) -> dict:
    d: dict = {
        "id":              p.id,
        "nama_promo":      p.nama_promo,
        "deskripsi":       p.deskripsi or "",
        "jenis_promo":     p.jenis_promo,
        "status":          p.status,
        "periode_mulai":   p.periode_mulai.isoformat() if p.periode_mulai else None,
        "periode_selesai": p.periode_selesai.isoformat() if p.periode_selesai else None,
        "created_by":      p.created_by,
        "created_at":      _iso_utc(p.created_at),
        "peserta":         [_peserta_to_dict(pp) for pp in peserta_rows],
        "summary_peserta": p.summary_peserta or {"total_toko": 0, "per_cluster": {}, "estimasi_budget_total": 0},
    }
    # reward_config kolom menyatukan konfigurasi_promo (v1) dan reward_config
    # (v2/v3) — lihat _V1_JENIS_PROMO di atas untuk aturan rekonstruksi nama field.
    if p.jenis_promo in _V1_JENIS_PROMO:
        d["konfigurasi_promo"] = p.reward_config or {}
    else:
        d["reward_config"] = p.reward_config or {}
    if p.tipe_program is not None:
        d["tipe_program"] = p.tipe_program
    if p.activated_at is not None:
        d["activated_at"] = _iso_utc(p.activated_at)
    if p.completed_at is not None:
        d["completed_at"] = _iso_utc(p.completed_at)
    if p.cancelled_at is not None:
        d["cancelled_at"] = _iso_utc(p.cancelled_at)
    if p.alasan_batal is not None:
        d["alasan_batal"] = p.alasan_batal
    if p.final_summary is not None:
        d["final_summary"] = p.final_summary
    if p.final_achievements is not None:
        d["final_achievements"] = p.final_achievements
    if p.brand_selection_mode is not None:
        d["brand_selection_mode"] = p.brand_selection_mode
    if p.brands is not None:
        d["brands"] = p.brands
    return d


def _get_promos() -> list[dict]:
    if USE_SQLITE:
        db = SessionLocal()
        try:
            rows = db.query(PromoRow).all()
            out = [
                _promo_row_to_dict(
                    p,
                    db.query(PromoPesertaRow).filter_by(promo_id=p.id).order_by(PromoPesertaRow.id).all(),
                )
                for p in rows
            ]
            out += [dict(a.raw_json) for a in db.query(PromoArchiveRow).all()]
            out.sort(key=lambda x: x["id"])
            return out
        finally:
            db.close()
    with _LOCK:
        return _rp(_PROMOS_PATH)


def _get_promo_by_id(promo_id: str) -> dict | None:
    if USE_SQLITE:
        db = SessionLocal()
        try:
            p = db.query(PromoRow).filter_by(id=promo_id).first()
            if p is not None:
                peserta_rows = db.query(PromoPesertaRow).filter_by(promo_id=p.id).order_by(PromoPesertaRow.id).all()
                return _promo_row_to_dict(p, peserta_rows)
            a = db.query(PromoArchiveRow).filter_by(id=promo_id).first()
            return dict(a.raw_json) if a is not None else None
        finally:
            db.close()
    promos = _rp(_PROMOS_PATH)
    return next((p for p in promos if p["id"] == promo_id), None)


def _save_new_promo(promo: dict) -> None:
    """Insert promo baru — selalu Draft + peserta kosong di kedua mode
    (satu-satunya jalur insert; lihat create/create-v2/create-v3)."""
    if USE_SQLITE:
        db = SessionLocal()
        try:
            reward_config = promo.get("konfigurasi_promo") if promo.get("jenis_promo") in _V1_JENIS_PROMO else promo.get("reward_config")
            db.add(PromoRow(
                id=promo["id"], nama_promo=promo["nama_promo"], deskripsi=promo.get("deskripsi", ""),
                jenis_promo=promo.get("jenis_promo"), tipe_program=promo.get("tipe_program"),
                status=promo["status"],
                periode_mulai=date.fromisoformat(promo["periode_mulai"]),
                periode_selesai=date.fromisoformat(promo["periode_selesai"]),
                created_by=promo.get("created_by", "admin"),
                reward_config=reward_config or {},
                summary_peserta=promo.get("summary_peserta", {}),
                brand_selection_mode=promo.get("brand_selection_mode"),
                brands=promo.get("brands"),
            ))
            db.commit()
        finally:
            db.close()
        return
    with _LOCK:
        promos = _rp(_PROMOS_PATH)
        promos.append(promo)
        _wp(_PROMOS_PATH, promos)


_DATETIME_FIELDS = {"activated_at", "completed_at", "cancelled_at"}


def _update_promo_status(promo_id: str, **fields: Any) -> dict | None:
    """Update field skalar pada Promo row (status/activated_at/dst). Hanya
    dipanggil utk promo di tabel Promo (status-guard di endpoint mencegah
    operasi ini menyentuh entri PromoArchive — lihat audit Tahap 4b).
    fields: nilai dalam format JSON-friendly (string ISO utk datetime,
    sama seperti _now()) — dikonversi ke tipe kolom di cabang SQLite."""
    if USE_SQLITE:
        db = SessionLocal()
        try:
            p = db.query(PromoRow).filter_by(id=promo_id).first()
            if p is None:
                return None
            for k, v in fields.items():
                if k in _DATETIME_FIELDS and isinstance(v, str):
                    v = datetime.fromisoformat(v)
                setattr(p, k, v)
            db.commit()
            peserta_rows = db.query(PromoPesertaRow).filter_by(promo_id=p.id).order_by(PromoPesertaRow.id).all()
            return _promo_row_to_dict(p, peserta_rows)
        finally:
            db.close()
    with _LOCK:
        promos = _rp(_PROMOS_PATH)
        idx = next((i for i, p in enumerate(promos) if p["id"] == promo_id), None)
        if idx is None:
            return None
        promos[idx].update(fields)
        _wp(_PROMOS_PATH, promos)
        return promos[idx]


def _delete_promo_by_id(promo_id: str) -> bool:
    if USE_SQLITE:
        db = SessionLocal()
        try:
            p = db.query(PromoRow).filter_by(id=promo_id).first()
            if p is not None:
                db.delete(p)  # cascades to PromoPeserta
                db.commit()
                return True
            a = db.query(PromoArchiveRow).filter_by(id=promo_id).first()
            if a is not None:
                db.delete(a)
                db.commit()
                return True
            return False
        finally:
            db.close()
    with _LOCK:
        promos = _rp(_PROMOS_PATH)
        idx = next((i for i, p in enumerate(promos) if p["id"] == promo_id), None)
        if idx is None:
            return False
        promos.pop(idx)
        _wp(_PROMOS_PATH, promos)
        return True


def _set_peserta_list(promo_id: str, peserta: list[dict], summary_peserta: dict) -> None:
    """Replace seluruh list peserta suatu promo (dipakai add-one/upload-excel/
    add/update/remove — semuanya menulis ulang list lengkap, sama seperti
    cabang JSON yang menulis ulang seluruh array setiap kali)."""
    if USE_SQLITE:
        db = SessionLocal()
        try:
            db.query(PromoPesertaRow).filter_by(promo_id=promo_id).delete()
            for p in peserta:
                db.add(PromoPesertaRow(
                    promo_id=promo_id, id_toko=p["id_toko"], nama_toko=p["nama_toko"],
                    cluster=p["cluster"], target_ton=p.get("target_ton") or 0.0,
                    rate_override=p.get("rate_override"), brand_utama=p.get("brand_utama"),
                    catatan=p.get("catatan", ""),
                ))
            row = db.query(PromoRow).filter_by(id=promo_id).first()
            if row is not None:
                row.summary_peserta = summary_peserta
            db.commit()
        finally:
            db.close()
        return
    with _LOCK:
        promos = _rp(_PROMOS_PATH)
        idx = next((i for i, p in enumerate(promos) if p["id"] == promo_id), None)
        if idx is None:
            return
        promos[idx]["peserta"]         = peserta
        promos[idx]["summary_peserta"] = summary_peserta
        _wp(_PROMOS_PATH, promos)


def _member_row_to_dict(m: LoyaltyMemberRow) -> dict:
    return {
        "id":             m.id,
        "id_toko":        m.id_toko,
        "nama_toko":      m.nama_toko,
        "kabupaten":      m.kabupaten,
        "cluster_pareto": m.cluster_pareto,
        "tso":            m.tso,
        "reward_type":    m.reward_type,
        "catatan":        m.catatan or "",
        "status":         m.status,
        "tgl_masuk":      m.tgl_masuk.isoformat() if m.tgl_masuk else None,
        "tgl_keluar":     m.tgl_keluar.isoformat() if m.tgl_keluar else None,
        "alasan_keluar":  m.alasan_keluar,
    }


def _get_loyalty_members_for_promo() -> list[dict]:
    """Cross-module lookup — promo.py butuh data toko loyalty utk validasi
    peserta (nama_toko/cluster/status). Mode-aware terhadap flag yang SAMA
    dgn loyalty.py: saat SQLite aktif, loyalty.py berhenti menulis ke
    loyalty_members.json sama sekali, jadi baca JSON di sini akan stale
    (lihat audit Tahap 4b) — wajib ikut baca dari SQLite saat flag true."""
    if USE_SQLITE:
        db = SessionLocal()
        try:
            return [_member_row_to_dict(m) for m in db.query(LoyaltyMemberRow).all()]
        finally:
            db.close()
    return _rp(_MEMBERS_PATH)


# ── Business helpers ──────────────────────────────────────────────────────────

def _generate_id(existing: list[dict]) -> str:
    today  = date.today().strftime("%Y%m%d")
    prefix = f"PROMO-{today}-"
    n      = sum(1 for p in existing if p["id"].startswith(prefix)) + 1
    return f"{prefix}{n:03d}"


def _infer_jenis(cfg: dict) -> str:
    rr  = cfg.get("reward_rate",  {}).get("enabled", False)
    tb  = cfg.get("target_bonus", {}).get("enabled", False)
    cb  = cfg.get("cashback",     {}).get("enabled", False)
    n   = sum([rr, tb, cb])
    if n > 1:  return "kombinasi"
    if rr:     return "reward_rate"
    if tb:     return "target_bonus"
    if cb:     return "cashback"
    return "reward_rate"


def _rebuild_summary(peserta: list[dict], konfigurasi: dict) -> dict:
    per_cluster: dict[str, int] = {}
    for p in peserta:
        cl = str(p.get("cluster", "Unknown"))
        per_cluster[cl] = per_cluster.get(cl, 0) + 1
    return {
        "total_toko":            len(peserta),
        "per_cluster":           per_cluster,
        "estimasi_budget_total": estimate_budget(peserta, konfigurasi),
    }


# ── Pydantic models ───────────────────────────────────────────────────────────

class RewardRateCfg(BaseModel):
    enabled:           bool  = False
    mode:              str   = "flat"
    flat_rate:         float = 10000
    per_cluster_rates: dict  = {}


class TargetBonusCfg(BaseModel):
    enabled:       bool  = False
    threshold_pct: float = 100
    bonus_rate:    float = 2000


class CashbackCfg(BaseModel):
    enabled:      bool  = False
    cashback_pct: float = 2.0


class KonfigurasiPromo(BaseModel):
    reward_rate:  RewardRateCfg  = RewardRateCfg()
    target_bonus: TargetBonusCfg = TargetBonusCfg()
    cashback:     CashbackCfg    = CashbackCfg()


class CreatePromoBody(BaseModel):
    nama_promo:        str
    deskripsi:         str              = ""
    periode_mulai:     str
    periode_selesai:   str
    konfigurasi_promo: KonfigurasiPromo = KonfigurasiPromo()
    created_by:        str              = "admin"


class UpdatePromoBody(BaseModel):
    nama_promo:        str | None              = None
    deskripsi:         str | None              = None
    periode_mulai:     str | None              = None
    periode_selesai:   str | None              = None
    konfigurasi_promo: KonfigurasiPromo | None = None


class CancelBody(BaseModel):
    alasan: str = ""


class AddPesertaBody(BaseModel):
    id_toko:       str
    target_ton:    float | None = None
    rate_override: float | None = None
    catatan:       str          = ""


class AddPesertaMonBody(BaseModel):
    """Add peserta from Monitoring tab — works for Draft and Aktif programs."""
    id_toko:     str
    nama_toko:   str | None = None
    cluster:     str | None = None
    target_ton:  float      = 0.0
    brand_utama: str | None = None
    catatan:     str        = ""


class UpdatePesertaBody(BaseModel):
    target_ton:  float | None = None
    brand_utama: str | None   = None
    catatan:     str | None   = None


# ── Multi-tier reward models ──────────────────────────────────────────────────

class TierConfig(BaseModel):
    tier_id:       int
    label:         str
    threshold_pct: float
    multiplier:    float
    keterangan:    str = ""


class RewardConfigMultiTier(BaseModel):
    type:                str        = "multi_tier_points"
    tiers:               list[TierConfig]
    reguler_multiplier:  float      = 1.0
    overflow_multiplier: float      = 1.0
    catatan:             str        = ""


class CreatePromoBodyV2(BaseModel):
    nama_promo:       str
    deskripsi:        str               = ""
    periode_mulai:    str
    periode_selesai:  str
    reward_config:    RewardConfigMultiTier
    created_by:       str               = "admin"


class UpdatePromoBodyV2(BaseModel):
    nama_promo:      str | None               = None
    deskripsi:       str | None               = None
    periode_mulai:   str | None               = None
    periode_selesai: str | None               = None
    reward_config:   RewardConfigMultiTier | None = None


class PreviewCalcBody(BaseModel):
    volume_realisasi: float
    volume_target:    float
    brand:            str
    tiers: list[TierConfig]
    reguler_multiplier:  float = 1.0
    overflow_multiplier: float = 1.0


# ── Unified (v3) models — 3 tipe program ─────────────────────────────────────

class LeaderboardRankReward(BaseModel):
    rank:         int | None       = None
    rank_range:   list[int] | None = None  # [lo, hi] inclusive
    label:        str
    reward_value: float


class CreatePromoV3Body(BaseModel):
    nama_promo:       str
    deskripsi:        str            = ""
    periode_mulai:    str
    periode_selesai:  str
    tipe_program:     str            # "flat_multiplier" | "multi_tier" | "leaderboard"
    reward_config:    dict[str, Any]
    created_by:       str            = "admin"
    brand_selection_mode: str | None = None  # "wilayah" | "fighting"
    brands:               list[dict[str, Any]] | None = None  # [{id, nama, tipe}], lihat brand_config_engine


# ── POST /api/promo/preview-calc ─────────────────────────────────────────────
# Static routes — must be registered before /{promo_id}

@router.post("/preview-calc")
def preview_calc(body: PreviewCalcBody) -> dict:
    """Kalkulasi reward langsung dari tiers tanpa simpan ke DB (untuk form preview)."""
    bpv = pc.get_brand_point_values()
    reward_config = {
        "tiers": [t.model_dump() for t in body.tiers],
        "reguler_multiplier":  body.reguler_multiplier,
        "overflow_multiplier": body.overflow_multiplier,
    }
    result = pc.calculate_tier_reward(
        volume_realisasi   = body.volume_realisasi,
        volume_target      = body.volume_target,
        reward_config      = reward_config,
        brand_name         = body.brand,
        brand_point_values = bpv,
    )
    return {"status": "ok", "data": result, "meta": _meta()}


# ── POST /api/promo/create-v2 (multi-tier reward) ─────────────────────────────

@router.post("/create-v2", status_code=201)
def create_promo_v2(body: CreatePromoBodyV2) -> dict:
    """Buat program promo dengan reward_config multi-tier."""
    rc_dict = body.reward_config.model_dump()

    # Validate tiers ordering
    tiers = sorted(rc_dict["tiers"], key=lambda x: x["threshold_pct"])
    if not tiers:
        raise HTTPException(400, "Minimal 1 tier harus didefinisikan")

    thresholds   = [t["threshold_pct"] for t in tiers]
    multipliers  = [t["multiplier"]    for t in tiers]

    if len(set(thresholds)) != len(thresholds):
        raise HTTPException(400, "threshold_pct harus unik, tidak boleh duplikat")
    if any(t <= 0 for t in thresholds):
        raise HTTPException(400, "threshold_pct harus lebih dari 0")
    if any(m <= 1 for m in multipliers):
        raise HTTPException(400, "multiplier setiap tier harus lebih dari 1")
    if multipliers != sorted(multipliers):
        raise HTTPException(400, "multiplier harus naik mengikuti urutan threshold tier")

    def _build(new_id: str) -> dict:
        return {
            "id":             new_id,
            "nama_promo":     body.nama_promo,
            "deskripsi":      body.deskripsi,
            "jenis_promo":    "multi_tier_points",
            "status":         "Draft",
            "periode_mulai":  body.periode_mulai,
            "periode_selesai": body.periode_selesai,
            "created_by":     body.created_by,
            "created_at":     _now(),
            "reward_config":  rc_dict,
            "peserta":        [],
            "summary_peserta": {"total_toko": 0, "per_cluster": {}, "estimasi_budget_total": 0},
        }

    if USE_SQLITE:
        promo = _build(_generate_id(_get_promos()))
        _save_new_promo(promo)
    else:
        with _LOCK:
            promos = _rp(_PROMOS_PATH)
            promo  = _build(_generate_id(promos))
            promos.append(promo)
            _wp(_PROMOS_PATH, promos)

    return {"status": "ok", "data": promo}


# ── POST /api/promo/create-v3 (unified 3-type) ───────────────────────────────

@router.post("/create-v3", status_code=201)
def create_promo_v3(body: CreatePromoV3Body) -> dict:
    """Buat program promo tipe 1 (flat_multiplier), 2 (multi_tier), atau 3 (leaderboard)."""
    VALID_TYPES = ("flat_multiplier", "multi_tier", "leaderboard")
    if body.tipe_program not in VALID_TYPES:
        raise HTTPException(400, f"tipe_program harus salah satu dari: {VALID_TYPES}")

    rc = body.reward_config

    if body.tipe_program == "flat_multiplier":
        mult = rc.get("multiplier", 0)
        if not isinstance(mult, (int, float)) or mult <= 1:
            raise HTTPException(400, "multiplier harus lebih dari 1")

    elif body.tipe_program == "multi_tier":
        tiers = rc.get("tiers", [])
        if not tiers:
            raise HTTPException(400, "Minimal 1 tier harus didefinisikan")
        thresholds  = [t["threshold_pct"] for t in tiers]
        multipliers = [t["multiplier"]    for t in tiers]
        if len(set(thresholds)) != len(thresholds):
            raise HTTPException(400, "threshold_pct harus unik")
        if any(m <= 1 for m in multipliers):
            raise HTTPException(400, "multiplier setiap tier harus lebih dari 1")

    elif body.tipe_program == "leaderboard":
        rr = rc.get("rank_rewards", [])
        if not rr:
            raise HTTPException(400, "Minimal 1 rank reward harus didefinisikan")
        if rc.get("basis_ranking") not in ("volume", "growth_pct"):
            raise HTTPException(400, "basis_ranking harus 'volume' atau 'growth_pct'")

    if body.brand_selection_mode is not None:
        if body.brand_selection_mode not in ("wilayah", "fighting"):
            raise HTTPException(400, "brand_selection_mode harus 'wilayah' atau 'fighting'")
        if not body.brands:
            raise HTTPException(400, "brands tidak boleh kosong kalau brand_selection_mode diisi")

        if body.tipe_program == "flat_multiplier":
            # brand_filter dipakai langsung oleh calculate_program_reward (promo_calculator.py)
            # untuk match dengan peserta["brand_utama"] — itu title-case ("Semen Elang"),
            # SEDANGKAN BrandConfig/resolve menyimpan UPPERCASE ("SEMEN ELANG", lihat
            # brand_config_engine.DEFAULT_CONFIG). title() di sini menjembatani dua
            # konvensi casing itu, bukan asumsi keduanya sama.
            rc["brand_filter"] = [str(b.get("nama", "")).title() for b in body.brands if b.get("nama")]

    jenis_map = {
        "flat_multiplier": "flat_multiplier",
        "multi_tier":      "multi_tier_points",
        "leaderboard":     "leaderboard",
    }

    def _build(new_id: str) -> dict:
        return {
            "id":              new_id,
            "nama_promo":      body.nama_promo,
            "deskripsi":       body.deskripsi,
            "jenis_promo":     jenis_map[body.tipe_program],
            "tipe_program":    body.tipe_program,
            "status":          "Draft",
            "periode_mulai":   body.periode_mulai,
            "periode_selesai": body.periode_selesai,
            "created_by":      body.created_by,
            "created_at":      _now(),
            "reward_config":   rc,
            "peserta":         [],
            "summary_peserta": {"total_toko": 0, "per_cluster": {}, "estimasi_budget_total": 0},
            "brand_selection_mode": body.brand_selection_mode,
            "brands":               body.brands,
        }

    if USE_SQLITE:
        promo = _build(_generate_id(_get_promos()))
        _save_new_promo(promo)
    else:
        with _LOCK:
            promos = _rp(_PROMOS_PATH)
            promo  = _build(_generate_id(promos))
            promos.append(promo)
            _wp(_PROMOS_PATH, promos)

    return {"status": "ok", "data": promo}


# ── GET /api/promo/template/peserta ──────────────────────────────────────────

@router.get("/template/peserta")
def download_template() -> StreamingResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Peserta"
    headers = ["ID Toko", "Target TON", "Rate Override", "Catatan"]
    ws.append(headers)
    ws.append(["TK001", 50.0, "", "Contoh peserta"])
    ws.append(["TK002", 45.0, 12000, "Rate override khusus"])

    hdr_fill = PatternFill("solid", fgColor="7C3AED")
    hdr_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = (
            max(len(str(c.value or "")) for c in col) + 4
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_peserta_promo.xlsx"},
    )


# ── GET /api/promo ────────────────────────────────────────────────────────────

@router.get("")
def list_promos(
    status: str | None = Query(None),
    limit:  int        = Query(50, ge=1, le=200),
    offset: int        = Query(0,  ge=0),
) -> dict:
    promos = _get_promos()

    if status:
        promos = [p for p in promos if p.get("status") == status]

    total = len(promos)
    page  = promos[offset: offset + limit]
    # Strip heavy arrays from list view
    slim  = [
        {k: v for k, v in p.items() if k not in ("peserta", "final_achievements")}
        for p in page
    ]
    return {"status": "ok", "data": slim, "meta": _meta(total=total, limit=limit, offset=offset)}


# ── POST /api/promo/create ────────────────────────────────────────────────────

@router.post("/create", status_code=201)
def create_promo(body: CreatePromoBody) -> dict:
    cfg  = body.konfigurasi_promo.model_dump()

    def _build(new_id: str) -> dict:
        return {
            "id":               new_id,
            "nama_promo":       body.nama_promo,
            "deskripsi":        body.deskripsi,
            "jenis_promo":      _infer_jenis(cfg),
            "status":           "Draft",
            "periode_mulai":    body.periode_mulai,
            "periode_selesai":  body.periode_selesai,
            "created_by":       body.created_by,
            "created_at":       _now(),
            "konfigurasi_promo": cfg,
            "peserta":          [],
            "summary_peserta":  {"total_toko": 0, "per_cluster": {}, "estimasi_budget_total": 0},
        }

    if USE_SQLITE:
        promo = _build(_generate_id(_get_promos()))
        _save_new_promo(promo)
    else:
        with _LOCK:
            promos = _rp(_PROMOS_PATH)
            promo  = _build(_generate_id(promos))
            promos.append(promo)
            _wp(_PROMOS_PATH, promos)

    return {"status": "ok", "data": promo}


# ── GET /api/promo/{promo_id} ─────────────────────────────────────────────────

@router.get("/{promo_id}")
def get_promo(promo_id: str) -> dict:
    promo = _get_promo_by_id(promo_id)
    if not promo:
        raise HTTPException(404, detail="Promo tidak ditemukan")
    return {"status": "ok", "data": promo, "meta": _meta()}


# ── PUT /api/promo/{promo_id}/update ─────────────────────────────────────────

@router.put("/{promo_id}/update")
def update_promo(promo_id: str, body: UpdatePromoBody) -> dict:
    if USE_SQLITE:
        db = SessionLocal()
        try:
            p = db.query(PromoRow).filter_by(id=promo_id).first()
            if p is None:
                raise HTTPException(404, detail="Promo tidak ditemukan")
            if p.status != "Draft":
                raise HTTPException(400, detail="Hanya promo Draft yang bisa diubah")

            if body.nama_promo is not None:
                p.nama_promo = body.nama_promo
            if body.deskripsi is not None:
                p.deskripsi = body.deskripsi
            if body.periode_mulai is not None:
                p.periode_mulai = date.fromisoformat(body.periode_mulai)
            if body.periode_selesai is not None:
                p.periode_selesai = date.fromisoformat(body.periode_selesai)
            peserta_rows = db.query(PromoPesertaRow).filter_by(promo_id=p.id).order_by(PromoPesertaRow.id).all()
            if body.konfigurasi_promo is not None:
                cfg = body.konfigurasi_promo.model_dump()
                p.reward_config   = cfg
                p.jenis_promo     = _infer_jenis(cfg)
                peserta_dicts     = [_peserta_to_dict(pp) for pp in peserta_rows]
                p.summary_peserta = _rebuild_summary(peserta_dicts, cfg)

            db.commit()
            updated = _promo_row_to_dict(p, peserta_rows)
        finally:
            db.close()
        return {"status": "ok", "data": updated}

    with _LOCK:
        promos = _rp(_PROMOS_PATH)
        idx = next((i for i, p in enumerate(promos) if p["id"] == promo_id), None)
        if idx is None:
            raise HTTPException(404, detail="Promo tidak ditemukan")
        if promos[idx]["status"] != "Draft":
            raise HTTPException(400, detail="Hanya promo Draft yang bisa diubah")

        if body.nama_promo is not None:
            promos[idx]["nama_promo"]    = body.nama_promo
        if body.deskripsi is not None:
            promos[idx]["deskripsi"]     = body.deskripsi
        if body.periode_mulai is not None:
            promos[idx]["periode_mulai"] = body.periode_mulai
        if body.periode_selesai is not None:
            promos[idx]["periode_selesai"] = body.periode_selesai
        if body.konfigurasi_promo is not None:
            cfg = body.konfigurasi_promo.model_dump()
            promos[idx]["konfigurasi_promo"] = cfg
            promos[idx]["jenis_promo"]       = _infer_jenis(cfg)
            promos[idx]["summary_peserta"]   = _rebuild_summary(promos[idx]["peserta"], cfg)

        _wp(_PROMOS_PATH, promos)
        updated = promos[idx]

    return {"status": "ok", "data": updated}


# ── POST /api/promo/{promo_id}/activate ──────────────────────────────────────

@router.post("/{promo_id}/activate")
def activate_promo(promo_id: str) -> dict:
    if USE_SQLITE:
        promo = _get_promo_by_id(promo_id)
        if promo is None:
            raise HTTPException(404, detail="Promo tidak ditemukan")
        if promo["status"] != "Draft":
            raise HTTPException(400, detail="Hanya promo Draft yang bisa diaktifkan")
        if not promo["peserta"]:
            raise HTTPException(400, detail="Promo harus memiliki minimal satu peserta")
        updated = _update_promo_status(promo_id, status="Aktif", activated_at=_now())
        return {"status": "ok", "data": updated}

    with _LOCK:
        promos = _rp(_PROMOS_PATH)
        idx = next((i for i, p in enumerate(promos) if p["id"] == promo_id), None)
        if idx is None:
            raise HTTPException(404, detail="Promo tidak ditemukan")
        if promos[idx]["status"] != "Draft":
            raise HTTPException(400, detail="Hanya promo Draft yang bisa diaktifkan")
        if not promos[idx]["peserta"]:
            raise HTTPException(400, detail="Promo harus memiliki minimal satu peserta")

        promos[idx]["status"]       = "Aktif"
        promos[idx]["activated_at"] = _now()
        _wp(_PROMOS_PATH, promos)
        updated = promos[idx]

    return {"status": "ok", "data": updated}


# ── POST /api/promo/{promo_id}/complete ──────────────────────────────────────

@router.post("/{promo_id}/complete")
def complete_promo(promo_id: str) -> dict:
    if USE_SQLITE:
        promo = _get_promo_by_id(promo_id)
        if promo is None:
            raise HTTPException(404, detail="Promo tidak ditemukan")
        if promo["status"] != "Aktif":
            raise HTTPException(400, detail="Hanya promo Aktif yang bisa diselesaikan")

        df_trx  = load_data()
        ach_df  = calculate_promo_achievement(promo, df_trx)
        summary = get_promo_summary(promo, ach_df)

        updated = _update_promo_status(
            promo_id,
            status="Selesai",
            completed_at=_now(),
            final_summary=summary,
            final_achievements=ach_df.to_dict("records") if not ach_df.empty else [],
        )
        return {"status": "ok", "data": updated, "meta": _meta(summary=summary)}

    with _LOCK:
        promos = _rp(_PROMOS_PATH)
        idx = next((i for i, p in enumerate(promos) if p["id"] == promo_id), None)
        if idx is None:
            raise HTTPException(404, detail="Promo tidak ditemukan")
        if promos[idx]["status"] != "Aktif":
            raise HTTPException(400, detail="Hanya promo Aktif yang bisa diselesaikan")
        promo = promos[idx]

    df_trx  = load_data()
    ach_df  = calculate_promo_achievement(promo, df_trx)
    summary = get_promo_summary(promo, ach_df)

    with _LOCK:
        promos = _rp(_PROMOS_PATH)
        idx = next((i for i, p in enumerate(promos) if p["id"] == promo_id), None)
        promos[idx]["status"]             = "Selesai"
        promos[idx]["completed_at"]       = _now()
        promos[idx]["final_summary"]      = summary
        promos[idx]["final_achievements"] = ach_df.to_dict("records") if not ach_df.empty else []
        _wp(_PROMOS_PATH, promos)
        updated = promos[idx]

    return {"status": "ok", "data": updated, "meta": _meta(summary=summary)}


# ── POST /api/promo/{promo_id}/cancel ────────────────────────────────────────

@router.post("/{promo_id}/cancel")
def cancel_promo(promo_id: str, body: CancelBody) -> dict:
    if USE_SQLITE:
        promo = _get_promo_by_id(promo_id)
        if promo is None:
            raise HTTPException(404, detail="Promo tidak ditemukan")
        if promo["status"] in ("Selesai", "Dibatalkan"):
            raise HTTPException(400, detail=f"Promo {promo['status']} tidak bisa dibatalkan")
        updated = _update_promo_status(promo_id, status="Dibatalkan", alasan_batal=body.alasan, cancelled_at=_now())
        return {"status": "ok", "data": updated}

    with _LOCK:
        promos = _rp(_PROMOS_PATH)
        idx = next((i for i, p in enumerate(promos) if p["id"] == promo_id), None)
        if idx is None:
            raise HTTPException(404, detail="Promo tidak ditemukan")
        if promos[idx]["status"] in ("Selesai", "Dibatalkan"):
            raise HTTPException(400, detail=f"Promo {promos[idx]['status']} tidak bisa dibatalkan")

        promos[idx]["status"]       = "Dibatalkan"
        promos[idx]["alasan_batal"] = body.alasan
        promos[idx]["cancelled_at"] = _now()
        _wp(_PROMOS_PATH, promos)
        updated = promos[idx]

    return {"status": "ok", "data": updated}


# ── DELETE /api/promo/{promo_id} ──────────────────────────────────────────────

@router.delete("/{promo_id}")
def delete_promo(promo_id: str) -> dict:
    if USE_SQLITE:
        if not _delete_promo_by_id(promo_id):
            raise HTTPException(404, detail="Promo tidak ditemukan")
        return {"status": "ok", "deleted_id": promo_id}

    with _LOCK:
        promos = _rp(_PROMOS_PATH)
        idx = next((i for i, p in enumerate(promos) if p["id"] == promo_id), None)
        if idx is None:
            raise HTTPException(404, detail="Promo tidak ditemukan")
        promos.pop(idx)
        _wp(_PROMOS_PATH, promos)
    return {"status": "ok", "deleted_id": promo_id}


# ── POST /api/promo/{promo_id}/peserta/add-one ───────────────────────────────

@router.post("/{promo_id}/peserta/add-one")
def add_peserta(promo_id: str, body: AddPesertaBody) -> dict:
    members_raw = _get_loyalty_members_for_promo()
    member      = next(
        (m for m in members_raw if str(m["id_toko"]) == body.id_toko and m.get("status") == "Aktif"),
        None,
    )
    if member:
        nama_toko = str(member["nama_toko"])
        cluster   = str(member["cluster_pareto"])
    else:
        crs     = get_store_crs()
        crs_row = crs[crs["ID Toko"].astype(str) == body.id_toko]
        if crs_row.empty:
            raise HTTPException(404, detail=f"Toko {body.id_toko} tidak ditemukan")
        row       = crs_row.iloc[0]
        nama_toko = str(row.get("Nama Toko", ""))
        cluster   = str(row.get("Cluster Pareto", "Bronze"))

    new_entry = {
        "id_toko":       body.id_toko,
        "nama_toko":     nama_toko,
        "cluster":       cluster,
        "rate_override": body.rate_override,
        "target_ton":    body.target_ton or 0.0,
        "catatan":       body.catatan,
    }

    if USE_SQLITE:
        promo = _get_promo_by_id(promo_id)
        if promo is None:
            raise HTTPException(404, detail="Promo tidak ditemukan")
        if promo["status"] != "Draft":
            raise HTTPException(400, detail="Peserta hanya bisa ditambahkan saat status Draft")
        existing_ids = {str(p["id_toko"]) for p in promo["peserta"]}
        if body.id_toko in existing_ids:
            raise HTTPException(409, detail=f"Toko {body.id_toko} sudah terdaftar")
        updated_peserta = promo["peserta"] + [new_entry]
        summary = _rebuild_summary(updated_peserta, promo.get("konfigurasi_promo", {}))
        _set_peserta_list(promo_id, updated_peserta, summary)
        return {"status": "ok", "data": updated_peserta}

    with _LOCK:
        promos = _rp(_PROMOS_PATH)
        idx = next((i for i, p in enumerate(promos) if p["id"] == promo_id), None)
        if idx is None:
            raise HTTPException(404, detail="Promo tidak ditemukan")
        if promos[idx]["status"] != "Draft":
            raise HTTPException(400, detail="Peserta hanya bisa ditambahkan saat status Draft")

        existing_ids = {str(p["id_toko"]) for p in promos[idx]["peserta"]}
        if body.id_toko in existing_ids:
            raise HTTPException(409, detail=f"Toko {body.id_toko} sudah terdaftar")

        promos[idx]["peserta"].append(new_entry)
        promos[idx]["summary_peserta"] = _rebuild_summary(
            promos[idx]["peserta"], promos[idx].get("konfigurasi_promo", {})
        )
        _wp(_PROMOS_PATH, promos)
        updated_peserta = promos[idx]["peserta"]

    return {"status": "ok", "data": updated_peserta}


# ── POST /api/promo/{promo_id}/peserta/upload-excel ──────────────────────────

@router.post("/{promo_id}/peserta/upload-excel")
def upload_peserta(promo_id: str, file: UploadFile = File(...)) -> dict:
    fname = file.filename or ""
    if not fname.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, detail="File harus berformat .xlsx atau .xls")

    raw = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        raise HTTPException(400, detail=f"File tidak dapat dibaca: {exc}") from exc

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, detail="File kosong atau tidak ada baris data")

    header  = [str(c or "").strip() for c in rows[0]]
    col_idx = {name: i for i, name in enumerate(header)}

    if "ID Toko" not in col_idx:
        raise HTTPException(400, detail="Kolom 'ID Toko' tidak ditemukan di header")

    def gcol(row: tuple, name: str, default: str = "") -> str:
        i = col_idx.get(name)
        if i is None or i >= len(row):
            return default
        return str(row[i] or "").strip()

    members_raw = _get_loyalty_members_for_promo() if USE_SQLITE else _rp(_MEMBERS_PATH)
    member_map  = {str(m["id_toko"]): m for m in members_raw if m.get("status") == "Aktif"}
    crs         = get_store_crs()
    crs_idx     = crs.set_index("ID Toko") if "ID Toko" in crs.columns else pd.DataFrame()

    berhasil = 0
    duplikat = 0
    errors: list[str] = []
    new_peserta: list[dict] = []

    def _process_rows(existing_ids: set[str]) -> None:
        nonlocal berhasil, duplikat
        for row_num, row in enumerate(rows[1:], start=2):
            id_toko = gcol(row, "ID Toko")
            if not id_toko:
                errors.append(f"Baris {row_num}: ID Toko kosong")
                continue
            if id_toko in existing_ids:
                duplikat += 1
                continue

            try:
                target_ton = float(gcol(row, "Target TON", "0") or "0")
            except ValueError:
                target_ton = 0.0

            rate_str = gcol(row, "Rate Override", "")
            try:
                rate_override: float | None = float(rate_str) if rate_str else None
            except ValueError:
                rate_override = None

            catatan = gcol(row, "Catatan", "")

            if id_toko in member_map:
                m = member_map[id_toko]
                nama_toko = str(m["nama_toko"])
                cluster   = str(m["cluster_pareto"])
            elif not crs_idx.empty and id_toko in crs_idx.index:
                nama_toko = str(crs_idx.at[id_toko, "Nama Toko"] or "")
                cluster   = str(crs_idx.at[id_toko, "Cluster Pareto"] or "Bronze")
            else:
                errors.append(f"Baris {row_num} ({id_toko}): Toko tidak ditemukan di data")
                continue

            new_peserta.append({
                "id_toko":       id_toko,
                "nama_toko":     nama_toko,
                "cluster":       cluster,
                "rate_override": rate_override,
                "target_ton":    target_ton,
                "catatan":       catatan,
            })
            existing_ids.add(id_toko)
            berhasil += 1

    if USE_SQLITE:
        promo = _get_promo_by_id(promo_id)
        if promo is None:
            raise HTTPException(404, detail="Promo tidak ditemukan")
        if promo["status"] != "Draft":
            raise HTTPException(400, detail="Upload hanya bisa saat status Draft")

        _process_rows({str(p["id_toko"]) for p in promo["peserta"]})

        if new_peserta:
            updated_peserta = promo["peserta"] + new_peserta
            summary = _rebuild_summary(updated_peserta, promo.get("konfigurasi_promo", {}))
            _set_peserta_list(promo_id, updated_peserta, summary)

        return {
            "status": "ok",
            "data":   {"berhasil": berhasil, "duplikat": duplikat, "errors": errors},
            "meta":   _meta(),
        }

    with _LOCK:
        promos = _rp(_PROMOS_PATH)
        idx = next((i for i, p in enumerate(promos) if p["id"] == promo_id), None)
        if idx is None:
            raise HTTPException(404, detail="Promo tidak ditemukan")
        if promos[idx]["status"] != "Draft":
            raise HTTPException(400, detail="Upload hanya bisa saat status Draft")

        _process_rows({str(p["id_toko"]) for p in promos[idx]["peserta"]})

        if new_peserta:
            promos[idx]["peserta"].extend(new_peserta)
            promos[idx]["summary_peserta"] = _rebuild_summary(
                promos[idx]["peserta"], promos[idx].get("konfigurasi_promo", {})
            )
            _wp(_PROMOS_PATH, promos)

    return {
        "status": "ok",
        "data":   {"berhasil": berhasil, "duplikat": duplikat, "errors": errors},
        "meta":   _meta(),
    }


# ── GET /api/promo/{promo_id}/peserta/search-toko ────────────────────────────

@router.get("/{promo_id}/peserta/search-toko")
def search_toko_for_promo(
    promo_id: str,
    q: str = Query("", min_length=0),
) -> dict:
    """Cari toko yang belum jadi peserta program ini."""
    promo = _get_promo_by_id(promo_id)
    if not promo:
        raise HTTPException(404, detail="Promo tidak ditemukan")

    existing_ids = {str(p["id_toko"]) for p in promo.get("peserta", [])}
    members_raw  = _get_loyalty_members_for_promo()
    q_lower      = q.lower().strip()

    results: list[dict] = []
    for m in members_raw:
        if m.get("status") != "Aktif":
            continue
        id_t = str(m.get("id_toko", ""))
        if id_t in existing_ids:
            continue
        nama = str(m.get("nama_toko", "")).lower()
        if q_lower and q_lower not in id_t.lower() and q_lower not in nama:
            continue
        results.append({
            "id_toko":        id_t,
            "nama_toko":      str(m.get("nama_toko", "")),
            "cluster_pareto": str(m.get("cluster_pareto", "Bronze")),
            "brand_utama":    str(m.get("brand_utama", "")),
        })
        if len(results) >= 20:
            break

    return {"status": "ok", "data": results, "meta": _meta()}


# ── POST /api/promo/{promo_id}/peserta/add ────────────────────────────────────

@router.post("/{promo_id}/peserta/add")
def add_peserta_mon(promo_id: str, body: AddPesertaMonBody) -> dict:
    """Tambah peserta (Draft atau Aktif)."""
    nama_toko = body.nama_toko
    cluster   = body.cluster

    if not nama_toko or not cluster:
        members_raw = _get_loyalty_members_for_promo()
        member = next((m for m in members_raw if str(m.get("id_toko", "")) == body.id_toko), None)
        if member:
            nama_toko = nama_toko or str(member["nama_toko"])
            cluster   = cluster   or str(member["cluster_pareto"])
        else:
            crs     = get_store_crs()
            crs_row = crs[crs["ID Toko"].astype(str) == body.id_toko]
            if crs_row.empty:
                raise HTTPException(404, detail=f"Toko {body.id_toko} tidak ditemukan")
            row       = crs_row.iloc[0]
            nama_toko = nama_toko or str(row.get("Nama Toko", ""))
            cluster   = cluster   or str(row.get("Cluster Pareto", "Bronze"))

    new_p = {
        "id_toko":     body.id_toko,
        "nama_toko":   nama_toko or "",
        "cluster":     cluster or "Bronze",
        "target_ton":  body.target_ton,
        "brand_utama": body.brand_utama or "",
        "catatan":     body.catatan,
    }

    if USE_SQLITE:
        promo = _get_promo_by_id(promo_id)
        if promo is None:
            raise HTTPException(404, detail="Promo tidak ditemukan")
        if promo["status"] in ("Selesai", "Dibatalkan"):
            raise HTTPException(400, detail="Tidak bisa menambah peserta ke program yang sudah Selesai atau Dibatalkan")
        if any(str(p["id_toko"]) == body.id_toko for p in promo["peserta"]):
            raise HTTPException(409, detail=f"Toko {body.id_toko} sudah terdaftar dalam program ini")
        updated_peserta = promo["peserta"] + [new_p]
        summary = _rebuild_summary(updated_peserta, promo.get("konfigurasi_promo", {}))
        _set_peserta_list(promo_id, updated_peserta, summary)
        return {"status": "ok", "data": new_p, "meta": _meta()}

    with _LOCK:
        promos = _rp(_PROMOS_PATH)
        idx = next((i for i, p in enumerate(promos) if p["id"] == promo_id), None)
        if idx is None:
            raise HTTPException(404, detail="Promo tidak ditemukan")
        if promos[idx]["status"] in ("Selesai", "Dibatalkan"):
            raise HTTPException(400, detail="Tidak bisa menambah peserta ke program yang sudah Selesai atau Dibatalkan")

        if any(str(p["id_toko"]) == body.id_toko for p in promos[idx]["peserta"]):
            raise HTTPException(409, detail=f"Toko {body.id_toko} sudah terdaftar dalam program ini")

        promos[idx]["peserta"].append(new_p)
        promos[idx]["summary_peserta"] = _rebuild_summary(
            promos[idx]["peserta"], promos[idx].get("konfigurasi_promo", {})
        )
        _wp(_PROMOS_PATH, promos)

    return {"status": "ok", "data": new_p, "meta": _meta()}


# ── PUT /api/promo/{promo_id}/peserta/{id_toko} ───────────────────────────────

@router.put("/{promo_id}/peserta/{id_toko}")
def update_peserta(promo_id: str, id_toko: str, body: UpdatePesertaBody) -> dict:
    """Update data peserta spesifik."""
    if USE_SQLITE:
        promo = _get_promo_by_id(promo_id)
        if promo is None:
            raise HTTPException(404, detail="Promo tidak ditemukan")
        if promo["status"] in ("Selesai", "Dibatalkan"):
            raise HTTPException(400, detail="Tidak bisa mengubah peserta di program yang sudah Selesai atau Dibatalkan")
        peserta = promo["peserta"]
        pidx = next((j for j, p in enumerate(peserta) if str(p["id_toko"]) == id_toko), None)
        if pidx is None:
            raise HTTPException(404, detail=f"Toko {id_toko} tidak ditemukan dalam program")

        if body.target_ton  is not None: peserta[pidx]["target_ton"]  = body.target_ton
        if body.brand_utama is not None: peserta[pidx]["brand_utama"] = body.brand_utama
        if body.catatan     is not None: peserta[pidx]["catatan"]     = body.catatan

        updated = peserta[pidx]
        summary = _rebuild_summary(peserta, promo.get("konfigurasi_promo", {}))
        _set_peserta_list(promo_id, peserta, summary)
        return {"status": "ok", "data": updated, "meta": _meta()}

    with _LOCK:
        promos = _rp(_PROMOS_PATH)
        idx = next((i for i, p in enumerate(promos) if p["id"] == promo_id), None)
        if idx is None:
            raise HTTPException(404, detail="Promo tidak ditemukan")
        if promos[idx]["status"] in ("Selesai", "Dibatalkan"):
            raise HTTPException(400, detail="Tidak bisa mengubah peserta di program yang sudah Selesai atau Dibatalkan")

        pidx = next(
            (j for j, p in enumerate(promos[idx]["peserta"]) if str(p["id_toko"]) == id_toko),
            None,
        )
        if pidx is None:
            raise HTTPException(404, detail=f"Toko {id_toko} tidak ditemukan dalam program")

        if body.target_ton  is not None: promos[idx]["peserta"][pidx]["target_ton"]  = body.target_ton
        if body.brand_utama is not None: promos[idx]["peserta"][pidx]["brand_utama"] = body.brand_utama
        if body.catatan     is not None: promos[idx]["peserta"][pidx]["catatan"]     = body.catatan

        updated = promos[idx]["peserta"][pidx]
        promos[idx]["summary_peserta"] = _rebuild_summary(
            promos[idx]["peserta"], promos[idx].get("konfigurasi_promo", {})
        )
        _wp(_PROMOS_PATH, promos)

    return {"status": "ok", "data": updated, "meta": _meta()}


# ── DELETE /api/promo/{promo_id}/peserta/{id_toko} ───────────────────────────

@router.delete("/{promo_id}/peserta/{id_toko}")
def remove_peserta(promo_id: str, id_toko: str) -> dict:
    if USE_SQLITE:
        promo = _get_promo_by_id(promo_id)
        if promo is None:
            raise HTTPException(404, detail="Promo tidak ditemukan")
        if promo["status"] in ("Selesai", "Dibatalkan"):
            raise HTTPException(400, detail="Tidak bisa menghapus peserta dari program yang sudah Selesai atau Dibatalkan")
        before = len(promo["peserta"])
        updated_peserta = [p for p in promo["peserta"] if str(p["id_toko"]) != id_toko]
        if len(updated_peserta) == before:
            raise HTTPException(404, detail=f"Toko {id_toko} tidak ditemukan di peserta")
        summary = _rebuild_summary(updated_peserta, promo.get("konfigurasi_promo", {}))
        _set_peserta_list(promo_id, updated_peserta, summary)
        return {"status": "ok", "data": updated_peserta}

    with _LOCK:
        promos = _rp(_PROMOS_PATH)
        idx = next((i for i, p in enumerate(promos) if p["id"] == promo_id), None)
        if idx is None:
            raise HTTPException(404, detail="Promo tidak ditemukan")
        if promos[idx]["status"] in ("Selesai", "Dibatalkan"):
            raise HTTPException(400, detail="Tidak bisa menghapus peserta dari program yang sudah Selesai atau Dibatalkan")

        before = len(promos[idx]["peserta"])
        promos[idx]["peserta"] = [p for p in promos[idx]["peserta"] if str(p["id_toko"]) != id_toko]
        if len(promos[idx]["peserta"]) == before:
            raise HTTPException(404, detail=f"Toko {id_toko} tidak ditemukan di peserta")

        cfg = promos[idx].get("konfigurasi_promo", {})
        promos[idx]["summary_peserta"] = _rebuild_summary(promos[idx]["peserta"], cfg)
        _wp(_PROMOS_PATH, promos)
        updated = promos[idx]["peserta"]

    return {"status": "ok", "data": updated}


# ── GET /api/promo/{promo_id}/reward-preview ─────────────────────────────────

@router.get("/{promo_id}/reward-preview")
def reward_preview(
    promo_id:         str,
    volume_realisasi: float = 0.0,
    volume_target:    float = 100.0,
    brand:            str   = "Semen Elang",
) -> dict:
    """Preview kalkulasi reward menggunakan reward_config program yang tersimpan."""
    promo = _get_promo_by_id(promo_id)
    if not promo:
        raise HTTPException(404, "Promo tidak ditemukan")

    reward_config = promo.get("reward_config")
    if not reward_config:
        raise HTTPException(400, "Program ini menggunakan konfigurasi lama (non-multi-tier)")

    bpv    = pc.get_brand_point_values()
    result = pc.calculate_tier_reward(
        volume_realisasi   = volume_realisasi,
        volume_target      = volume_target,
        reward_config      = reward_config,
        brand_name         = brand,
        brand_point_values = bpv,
    )
    return {"status": "ok", "data": result, "meta": _meta()}


# ── GET /api/promo/{promo_id}/monitoring ─────────────────────────────────────

@router.get("/{promo_id}/monitoring")
def get_monitoring(
    promo_id: str,
    sort_by:  str = Query("achievement", pattern="^(achievement|realisasi|reward)$"),
    order:    str = Query("desc",        pattern="^(asc|desc)$"),
) -> dict:
    promo = _get_promo_by_id(promo_id)
    if not promo:
        raise HTTPException(404, detail="Promo tidak ditemukan")
    if promo["status"] not in ("Aktif", "Selesai"):
        raise HTTPException(400, detail="Monitoring hanya tersedia untuk promo Aktif atau Selesai")

    df_trx = load_data()

    # Detect tipe_program
    tipe = promo.get("tipe_program") or (
        "multi_tier" if promo.get("reward_config") else "legacy"
    )

    if tipe in ("flat_multiplier", "multi_tier", "leaderboard"):
        loyalty_cfg = pc.load_loyalty_config()
        result = pc.calculate_program_reward(
            promo         = promo,
            peserta_data  = promo.get("peserta", []),
            transaksi_df  = df_trx,
            loyalty_config = loyalty_cfg,
        )
        # backward-compat flag for multi_tier
        result["is_multi_tier"] = (tipe == "multi_tier")
        return {"status": "ok", "data": result, "meta": _meta()}

    # Legacy konfigurasi_promo flow
    if promo["status"] == "Selesai" and "final_achievements" in promo:
        ach_df = pd.DataFrame(promo["final_achievements"])
    else:
        ach_df = calculate_promo_achievement(promo, df_trx)
    summary = get_promo_summary(promo, ach_df)

    daily_trend = get_daily_trend(promo, df_trx)
    cluster_cmp = get_cluster_comparison(promo, df_trx) if promo["status"] == "Selesai" else []

    # Sort achievements
    sort_col = {"achievement": "achievement_pct", "realisasi": "realisasi_ton", "reward": "total_reward"}
    asc      = order == "asc"
    if not ach_df.empty:
        ach_df = ach_df.sort_values(sort_col.get(sort_by, "achievement_pct"), ascending=asc)

    # Distribution buckets
    if not ach_df.empty:
        pct = ach_df["achievement_pct"]
        distribution = [
            {"label": "0–50%",   "count": int((pct <= 50).sum()),                            "color": "#DC2626"},
            {"label": "51–70%",  "count": int(((pct > 50) & (pct <= 70)).sum()),              "color": "#EA580C"},
            {"label": "71–90%",  "count": int(((pct > 70) & (pct <= 90)).sum()),              "color": "#D97706"},
            {"label": "91–100%", "count": int(((pct > 90) & (pct <= 100)).sum()),             "color": "#16a34a"},
            {"label": ">100%",   "count": int((pct > 100).sum()),                             "color": "#059669"},
        ]
    else:
        distribution = []

    # Auto-recommendations for Selesai
    recommendations: list[str] = []
    if promo["status"] == "Selesai" and not ach_df.empty:
        melampaui = int((ach_df["achievement_pct"] > 100).sum())
        tidak_trx = int((ach_df["realisasi_ton"] == 0).sum())
        low_ach   = int((ach_df["achievement_pct"] < 80).sum())
        if melampaui > 0:
            recommendations.append(
                f"{melampaui} toko melampaui target — kandidat kuat untuk program promo berikutnya."
            )
        if tidak_trx > 0:
            recommendations.append(
                f"{tidak_trx} toko tidak bertransaksi selama periode promo — perlu evaluasi relevansi program."
            )
        if low_ach > 0:
            recommendations.append(
                f"{low_ach} toko achievement < 80% — pertimbangkan follow-up atau penyesuaian target."
            )

    return {
        "status": "ok",
        "data": {
            "tipe_program":       "legacy",
            "is_multi_tier":      False,
            "achievements":       ach_df.where(pd.notna(ach_df), None).to_dict("records") if not ach_df.empty else [],
            "summary":            summary,
            "daily_trend":        daily_trend,
            "distribution":       distribution,
            "cluster_comparison": cluster_cmp,
            "recommendations":    recommendations,
        },
        "meta": _meta(),
    }


# ── GET /api/promo/{promo_id}/standings ──────────────────────────────────────

@router.get("/{promo_id}/standings")
def get_standings(promo_id: str) -> dict:
    """Real-time leaderboard standings — khusus tipe_program leaderboard."""
    promo = _get_promo_by_id(promo_id)
    if not promo:
        raise HTTPException(404, detail="Promo tidak ditemukan")
    if promo.get("tipe_program") != "leaderboard":
        raise HTTPException(400, detail="Endpoint ini hanya untuk program tipe leaderboard")
    if promo["status"] not in ("Aktif", "Selesai"):
        raise HTTPException(400, detail="Standings hanya tersedia saat Aktif atau Selesai")

    df_trx      = load_data()
    loyalty_cfg = pc.load_loyalty_config()
    result      = pc.calculate_leaderboard_standings(
        promo         = promo,
        peserta_data  = promo.get("peserta", []),
        transaksi_df  = df_trx,
        loyalty_config = loyalty_cfg,
    )
    return {"status": "ok", "data": result, "meta": _meta()}


# ── GET /api/promo/{promo_id}/monitoring/export ──────────────────────────────

@router.get("/{promo_id}/monitoring/export")
def export_monitoring(promo_id: str) -> StreamingResponse:
    promo = _get_promo_by_id(promo_id)
    if not promo:
        raise HTTPException(404, detail="Promo tidak ditemukan")
    if promo["status"] not in ("Aktif", "Selesai"):
        raise HTTPException(400, detail="Export hanya untuk promo Aktif atau Selesai")

    df_trx = load_data()
    if promo["status"] == "Selesai" and "final_achievements" in promo:
        ach_df = pd.DataFrame(promo["final_achievements"])
    else:
        ach_df = calculate_promo_achievement(promo, df_trx)

    summary = get_promo_summary(promo, ach_df)

    wb = openpyxl.Workbook()

    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["LAPORAN MONITORING PROMO"])
    ws1.append([promo["nama_promo"]])
    ws1.append(["Periode", f"{promo['periode_mulai']} s/d {promo['periode_selesai']}"])
    ws1.append(["Status", promo["status"]])
    ws1.append([])
    ws1.append(["Total Peserta", summary["total_peserta"]])
    ws1.append(["Peserta Bertransaksi", summary["peserta_aktif_transaksi"]])
    ws1.append(["Total Target TON", summary["total_target_ton"]])
    ws1.append(["Total Realisasi TON", summary["total_realisasi_ton"]])
    ws1.append(["Overall Achievement", f"{summary['overall_achievement_pct']:.1f}%"])
    ws1.append(["Total Reward Earned (Rp)", summary["total_reward_earned"]])
    ws1.append(["Estimasi Budget Sisa (Rp)", summary["estimasi_budget_sisa"]])

    # Sheet 2: Achievement per toko
    ws2 = wb.create_sheet("Achievement per Toko")
    hdr = ["ID Toko", "Nama Toko", "Cluster", "Target TON", "Realisasi TON",
           "Achievement %", "Reward Rate (Rp)", "Bonus (Rp)", "Cashback (Rp)",
           "Total Reward (Rp)", "Status"]
    ws2.append(hdr)

    hdr_fill = PatternFill("solid", fgColor="7C3AED")
    hdr_font = Font(bold=True, color="FFFFFF")
    for cell in ws2[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font

    if not ach_df.empty:
        for _, row in ach_df.sort_values("achievement_pct", ascending=False).iterrows():
            ws2.append([
                row["id_toko"], row["nama_toko"], row["cluster"],
                float(row["target_ton"]), float(row["realisasi_ton"]),
                round(float(row["achievement_pct"]), 2),
                int(row["reward_rate_earned"]), int(row["bonus_earned"]),
                int(row["cashback_earned"]), int(row["total_reward"]),
                row["status"],
            ])

    for col in ws2.columns:
        ws2.column_dimensions[col[0].column_letter].width = (
            max(len(str(c.value or "")) for c in col) + 3
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = promo["nama_promo"].replace(" ", "_")[:30]
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=monitoring_{safe}.xlsx"},
    )
