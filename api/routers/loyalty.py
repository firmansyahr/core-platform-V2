"""Loyalty program management endpoints."""
from __future__ import annotations

import io
import json
import os
import threading
import uuid
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
import pandas as pd
from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from api.core.aegis_engine import get_store_crs
from api.core.data_loader import load_data
from api.core.ilp_engine import get_ilp_features
from api.core.loyalty_engine import (
    DEFAULT_CONFIG,
    REWARD_RATES,
    calculate_effectiveness,
    calculate_historical_targets,
    calculate_loyalty_targets,
    get_growth_rate,
    get_ilp_recommendations,
    get_smart_promotions,
    get_takeout_recommendations,
    get_target_triggers,
)
from api.database import SessionLocal
from api.models import LoyaltyConfig as LoyaltyConfigRow
from api.models import LoyaltyHistory as LoyaltyHistoryRow
from api.models import LoyaltyMember as LoyaltyMemberRow

router = APIRouter(prefix="/api/loyalty", tags=["loyalty"])

# Feature flag — Tahap 4 rollout. Default false = perilaku identik dengan
# sebelumnya (JSON). Set USE_SQLITE_STORAGE=true di env untuk pakai SQLite.
# Lihat fungsi _get_members()/_add_member()/dst. di bawah untuk abstraksinya.
USE_SQLITE = os.getenv("USE_SQLITE_STORAGE", "false").lower() == "true"

_LOCK = threading.Lock()


def _get_data_dir() -> Path:
    vol = Path("/mnt/data")
    if vol.exists() and os.access(vol, os.W_OK):
        d = vol / "app_data"
        d.mkdir(parents=True, exist_ok=True)
        return d
    d = Path(__file__).parent.parent / "data"
    d.mkdir(parents=True, exist_ok=True)
    return d


_DATA_DIR     = _get_data_dir()
_MEMBERS_PATH = _DATA_DIR / "loyalty_members.json"
_HISTORY_PATH = _DATA_DIR / "loyalty_history.json"
_CONFIG_PATH  = _DATA_DIR / "loyalty_config.json"

VALID_REWARD_TYPES = set(REWARD_RATES.keys())
VALID_CLUSTERS = {"Super Platinum", "Platinum", "Gold", "Silver", "Bronze"}
VALID_STATUS   = {"Aktif", "Nonaktif"}


# ── File helpers ──────────────────────────────────────────────────────────────

def _rj(path: Path, default: Any = None) -> Any:
    if not path.exists():
        return default if default is not None else []
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default if default is not None else []


def _wj(path: Path, data: Any) -> None:
    _DATA_DIR.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


# ── Storage abstraction (JSON / SQLite) ────────────────────────────────────────
#
# Setiap fungsi di bawah punya DUA cabang (USE_SQLITE True/False) yang
# mengembalikan/menerima bentuk dict YANG SAMA PERSIS dengan JSON asli, supaya
# endpoint di bawahnya (dan frontend) tidak perlu tahu/berubah sama sekali.
# Cabang JSON adalah logic lama, TIDAK diubah — hanya dipindah ke dalam fungsi.

def _member_to_dict(m: LoyaltyMemberRow) -> dict:
    return {
        "id":             m.id,
        "id_toko":        m.id_toko,
        "nama_toko":      m.nama_toko,
        "kabupaten":      m.kabupaten,
        "cluster_pareto": m.cluster_pareto,
        "tso":            m.tso,
        "reward_type":    m.reward_type,
        "catatan":        m.catatan or "",
        "status":         m.status,
        "tgl_masuk":      m.tgl_masuk.isoformat() if m.tgl_masuk else None,
        "tgl_keluar":     m.tgl_keluar.isoformat() if m.tgl_keluar else None,
        "alasan_keluar":  m.alasan_keluar,
    }


def _get_members() -> list[dict]:
    if USE_SQLITE:
        db = SessionLocal()
        try:
            return [_member_to_dict(m) for m in db.query(LoyaltyMemberRow).all()]
        finally:
            db.close()
    with _LOCK:
        return _rj(_MEMBERS_PATH)


def _get_member_by_id(member_id: str) -> dict | None:
    if USE_SQLITE:
        db = SessionLocal()
        try:
            row = db.query(LoyaltyMemberRow).filter_by(id=member_id).first()
            return _member_to_dict(row) if row else None
        finally:
            db.close()
    members = _get_members()
    return next((m for m in members if m.get("id") == member_id), None)


def _get_active_ids() -> set[str]:
    if USE_SQLITE:
        db = SessionLocal()
        try:
            rows = db.query(LoyaltyMemberRow.id_toko).filter_by(status="Aktif").all()
            return {r[0] for r in rows}
        finally:
            db.close()
    members = _get_members()
    return {m["id_toko"] for m in members if m.get("status") == "Aktif"}


def _is_id_toko_active(id_toko: str) -> bool:
    if USE_SQLITE:
        db = SessionLocal()
        try:
            return db.query(LoyaltyMemberRow).filter_by(id_toko=id_toko, status="Aktif").first() is not None
        finally:
            db.close()
    members = _get_members()
    return any(m["id_toko"] == id_toko and m.get("status") == "Aktif" for m in members)


def _add_member(new_member: dict) -> None:
    if USE_SQLITE:
        db = SessionLocal()
        try:
            db.add(LoyaltyMemberRow(
                id=new_member["id"], id_toko=new_member["id_toko"],
                nama_toko=new_member["nama_toko"], kabupaten=new_member["kabupaten"],
                cluster_pareto=new_member["cluster_pareto"], tso=new_member["tso"],
                reward_type=new_member["reward_type"], catatan=new_member.get("catatan") or "",
                status=new_member["status"], tgl_masuk=date.fromisoformat(new_member["tgl_masuk"]),
                tgl_keluar=None, alasan_keluar=None,
            ))
            db.commit()
        finally:
            db.close()
        return
    with _LOCK:
        members = _rj(_MEMBERS_PATH)
        members.append(new_member)
        _wj(_MEMBERS_PATH, members)


def _add_member_if_not_duplicate(new_member: dict) -> bool:
    """Insert toko baru HANYA kalau id_toko belum aktif — check+insert atomic
    dalam satu critical section per mode (sama seperti logic asli add_one_member,
    yang menahan _LOCK untuk seluruh check+append+write). Return False kalau
    sudah ada toko aktif dengan id_toko itu (caller raise 409)."""
    if USE_SQLITE:
        db = SessionLocal()
        try:
            exists = db.query(LoyaltyMemberRow).filter_by(
                id_toko=new_member["id_toko"], status="Aktif"
            ).first()
            if exists:
                return False
            db.add(LoyaltyMemberRow(
                id=new_member["id"], id_toko=new_member["id_toko"],
                nama_toko=new_member["nama_toko"], kabupaten=new_member["kabupaten"],
                cluster_pareto=new_member["cluster_pareto"], tso=new_member["tso"],
                reward_type=new_member["reward_type"], catatan=new_member.get("catatan") or "",
                status=new_member["status"], tgl_masuk=date.fromisoformat(new_member["tgl_masuk"]),
                tgl_keluar=None, alasan_keluar=None,
            ))
            db.commit()
            return True
        finally:
            db.close()
    with _LOCK:
        members = _rj(_MEMBERS_PATH)
        if any(m["id_toko"] == new_member["id_toko"] and m.get("status") == "Aktif" for m in members):
            return False
        members.append(new_member)
        _wj(_MEMBERS_PATH, members)
        return True


def _add_members_bulk(new_members: list[dict]) -> None:
    if not new_members:
        return
    if USE_SQLITE:
        db = SessionLocal()
        try:
            for nm in new_members:
                db.add(LoyaltyMemberRow(
                    id=nm["id"], id_toko=nm["id_toko"], nama_toko=nm["nama_toko"],
                    kabupaten=nm["kabupaten"], cluster_pareto=nm["cluster_pareto"],
                    tso=nm["tso"], reward_type=nm["reward_type"], catatan=nm.get("catatan") or "",
                    status=nm["status"], tgl_masuk=date.fromisoformat(nm["tgl_masuk"]),
                    tgl_keluar=None, alasan_keluar=None,
                ))
            db.commit()
        finally:
            db.close()
        return
    with _LOCK:
        members = _rj(_MEMBERS_PATH)
        members.extend(new_members)
        _wj(_MEMBERS_PATH, members)


def _update_member_by_id(member_id: str, updates: dict) -> dict | None:
    """updates: dict field->value sudah dalam format JSON-friendly (tanggal
    sebagai string ISO atau None) — dikonversi ke tipe kolom di cabang SQLite."""
    if USE_SQLITE:
        db = SessionLocal()
        try:
            row = db.query(LoyaltyMemberRow).filter_by(id=member_id).first()
            if row is None:
                return None
            for k, v in updates.items():
                if k in ("tgl_keluar", "tgl_masuk") and isinstance(v, str):
                    v = date.fromisoformat(v)
                setattr(row, k, v)
            db.commit()
            return _member_to_dict(row)
        finally:
            db.close()
    with _LOCK:
        members = _rj(_MEMBERS_PATH)
        idx = next((i for i, m in enumerate(members) if m.get("id") == member_id), None)
        if idx is None:
            return None
        members[idx] = {**members[idx], **updates}
        _wj(_MEMBERS_PATH, members)
        return members[idx]


def _history_to_dict(h: LoyaltyHistoryRow) -> dict:
    d = {
        "id_member":   h.id_member,
        "id_toko":     h.id_toko,
        "nama_toko":   h.nama_toko,
        "tanggal":     h.tanggal.isoformat(),
        "perubahan":   h.perubahan,
        "alasan":      h.alasan,
        "status_baru": h.status_baru,
    }
    # "catatan" tidak selalu ada di JSON asli (hanya dari take-out) — pertahankan
    # nuansa itu: key absen kalau None, bukan present-tapi-null, untuk parity persis.
    if h.catatan is not None:
        d["catatan"] = h.catatan
    return d


def _get_history() -> list[dict]:
    if USE_SQLITE:
        db = SessionLocal()
        try:
            return [_history_to_dict(h) for h in db.query(LoyaltyHistoryRow).all()]
        finally:
            db.close()
    with _LOCK:
        return _rj(_HISTORY_PATH)


def _add_history_event(event: dict) -> None:
    if USE_SQLITE:
        db = SessionLocal()
        try:
            db.add(LoyaltyHistoryRow(
                id_member=event.get("id_member"), id_toko=event["id_toko"],
                nama_toko=event["nama_toko"], tanggal=datetime.fromisoformat(event["tanggal"]),
                perubahan=event["perubahan"], alasan=event.get("alasan"),
                catatan=event.get("catatan"), status_baru=event["status_baru"],
            ))
            db.commit()
        finally:
            db.close()
        return
    with _LOCK:
        hist = _rj(_HISTORY_PATH)
        hist.append(event)
        _wj(_HISTORY_PATH, hist)


def _add_history_events_bulk(events: list[dict]) -> None:
    if not events:
        return
    if USE_SQLITE:
        db = SessionLocal()
        try:
            for event in events:
                db.add(LoyaltyHistoryRow(
                    id_member=event.get("id_member"), id_toko=event["id_toko"],
                    nama_toko=event["nama_toko"], tanggal=datetime.fromisoformat(event["tanggal"]),
                    perubahan=event["perubahan"], alasan=event.get("alasan"),
                    catatan=event.get("catatan"), status_baru=event["status_baru"],
                ))
            db.commit()
        finally:
            db.close()
        return
    with _LOCK:
        hist = _rj(_HISTORY_PATH)
        hist.extend(events)
        _wj(_HISTORY_PATH, hist)


def _save_target_config(new_cfg: dict) -> None:
    if USE_SQLITE:
        db = SessionLocal()
        try:
            row = db.query(LoyaltyConfigRow).filter_by(id="default").first()
            if row is None:
                row = LoyaltyConfigRow(id="default")
                db.add(row)
            row.w1               = new_cfg["w1"]
            row.w2               = new_cfg["w2"]
            row.min_pct_sp       = new_cfg["min_pct_sp"]
            row.min_pct_platinum = new_cfg["min_pct_platinum"]
            row.min_pct_gold     = new_cfg["min_pct_gold"]
            row.min_pct_silver   = new_cfg["min_pct_silver"]
            row.growth_rates     = new_cfg["growth_rates"]
            db.commit()
        finally:
            db.close()
        return
    _wj(_CONFIG_PATH, new_cfg)


def _log(event: dict) -> None:
    _add_history_event(event)


def _meta(**kw: Any) -> dict:
    return {"generated_at": _now(), **kw}


def _enrich_member(m: dict, crs_idx: pd.DataFrame, ilp_idx: pd.DataFrame) -> dict:
    """Add aegis_score, aegis_level, avg_ton_bulanan, est_budget to member dict."""
    out = dict(m)
    id_toko = str(m["id_toko"])

    if not crs_idx.empty and id_toko in crs_idx.index:
        out["aegis_score"] = round(float(crs_idx.at[id_toko, "aegis_score"] or 0), 2)
        out["aegis_level"] = str(crs_idx.at[id_toko, "alert"] or "Normal")
    else:
        out["aegis_score"] = 0.0
        out["aegis_level"] = "Normal"

    avg_ton = 0.0
    if not ilp_idx.empty and id_toko in ilp_idx.index:
        avg_ton = float(ilp_idx.at[id_toko, "avg_ton"] or 0)
    rate = REWARD_RATES.get(str(m.get("reward_type", "Standard")), 5_000)
    out["avg_ton_bulanan"] = round(avg_ton, 2)
    out["est_budget"]      = round(avg_ton * rate)
    return out


# ── Pydantic models ───────────────────────────────────────────────────────────

class AddOneMember(BaseModel):
    id_toko:        str
    nama_toko:      str
    kabupaten:      str
    cluster_pareto: str
    tso:            str
    reward_type:    str = "Standard"
    catatan:        str = ""


class TakeOutBody(BaseModel):
    alasan:  str
    catatan: str = ""


class UpdateRewardTypeBody(BaseModel):
    reward_type: str


class GrowthDefaultCfg(BaseModel):
    normal:  float = 0.03
    warning: float = 0.01
    kritis:  float = 0.00


class GrowthOverrideCfg(BaseModel):
    id:      str        = ""
    label:   str
    type:    str                # "monthly" | "quarterly"
    bulan:   int | None = None
    tahun:   int
    kuartal: int | None = None
    normal:  float
    warning: float
    kritis:  float
    catatan: str = ""


class GrowthRatesCfg(BaseModel):
    default:   GrowthDefaultCfg        = GrowthDefaultCfg()
    overrides: list[GrowthOverrideCfg] = []


class TargetConfigBody(BaseModel):
    w1:               float          = 0.6
    w2:               float          = 0.4
    min_pct_sp:       float          = 0.80
    min_pct_platinum: float          = 0.70
    min_pct_gold:     float          = 0.60
    min_pct_silver:   float          = 0.50
    growth_rates:     GrowthRatesCfg = GrowthRatesCfg()


def _load_config() -> dict:
    """Load loyalty_config (JSON atau SQLite, lihat USE_SQLITE) merged dengan
    DEFAULT_CONFIG sebagai fallback.

    Cabang JSON memigrasikan format flat lama growth_normal/warning/kritis →
    nested growth_rates — logic asli, tidak diubah.
    """
    base = dict(DEFAULT_CONFIG)

    if USE_SQLITE:
        db = SessionLocal()
        try:
            row = db.query(LoyaltyConfigRow).filter_by(id="default").first()
            if row is None:
                return base
            stored = {
                "w1": row.w1, "w2": row.w2,
                "min_pct_sp": row.min_pct_sp, "min_pct_platinum": row.min_pct_platinum,
                "min_pct_gold": row.min_pct_gold, "min_pct_silver": row.min_pct_silver,
                "growth_rates": row.growth_rates,
            }
            return {**base, **stored}
        finally:
            db.close()

    if not _CONFIG_PATH.exists():
        return base
    try:
        stored = json.loads(_CONFIG_PATH.read_text(encoding="utf-8"))
        # Migrate old flat format → nested growth_rates
        if "growth_normal" in stored and "growth_rates" not in stored:
            stored["growth_rates"] = {
                "default": {
                    "normal":  stored.pop("growth_normal",  0.03),
                    "warning": stored.pop("growth_warning", 0.01),
                    "kritis":  stored.pop("growth_kritis",  0.00),
                },
                "overrides": [],
            }
        return {**base, **stored}
    except Exception:
        return base


# ── GET /api/loyalty/members ──────────────────────────────────────────────────

@router.get("/members")
def get_members(
    status:      str | None = Query(None),
    cluster:     str | None = Query(None),
    reward_type: str | None = Query(None),
    limit:       int        = Query(50, ge=1, le=500),
    offset:      int        = Query(0, ge=0),
) -> dict:
    members = _get_members()

    if status:
        members = [m for m in members if m.get("status") == status]
    if cluster:
        members = [m for m in members if m.get("cluster_pareto") == cluster]
    if reward_type:
        members = [m for m in members if m.get("reward_type") == reward_type]

    crs     = get_store_crs()
    crs_idx = crs.set_index("ID Toko") if "ID Toko" in crs.columns else pd.DataFrame()
    ilp     = get_ilp_features()
    ilp_idx = (
        ilp.set_index("ID Toko")
        if not ilp.empty and "ID Toko" in ilp.columns
        else pd.DataFrame()
    )

    total = len(members)
    page  = members[offset: offset + limit]
    data  = [_enrich_member(m, crs_idx, ilp_idx) for m in page]

    return {"status": "ok", "data": data, "meta": _meta(total=total, limit=limit, offset=offset)}


# ── GET /api/loyalty/summary ──────────────────────────────────────────────────

@router.get("/summary")
def get_summary() -> dict:
    members = _get_members()

    aktif    = [m for m in members if m.get("status") == "Aktif"]
    nonaktif = [m for m in members if m.get("status") == "Nonaktif"]

    crs     = get_store_crs()
    crs_idx = crs.set_index("ID Toko") if "ID Toko" in crs.columns else pd.DataFrame()
    ilp     = get_ilp_features()
    ilp_idx = (
        ilp.set_index("ID Toko")
        if not ilp.empty and "ID Toko" in ilp.columns
        else pd.DataFrame()
    )

    est_budget_total = 0.0
    per_reward_type: dict[str, int] = {}
    for m in aktif:
        id_toko = str(m["id_toko"])
        avg_ton = (
            float(ilp_idx.at[id_toko, "avg_ton"] or 0)
            if not ilp_idx.empty and id_toko in ilp_idx.index
            else 0.0
        )
        rate = REWARD_RATES.get(str(m.get("reward_type", "Standard")), 5_000)
        est_budget_total += avg_ton * rate
        rt = str(m.get("reward_type", "Standard"))
        per_reward_type[rt] = per_reward_type.get(rt, 0) + 1

    takeout_count = 0
    efektivitas_data: dict = {}
    if aktif:
        df_trx = load_data()
        recs = get_takeout_recommendations(pd.DataFrame(aktif), df_trx, crs)
        takeout_count = len(recs)
        cfg    = _load_config()
        targets_df = calculate_loyalty_targets(pd.DataFrame(aktif), df_trx, crs, config=cfg)
        if not targets_df.empty:
            latest_p = targets_df["bulan_target"].iloc[0]
            try:
                y, mo = str(latest_p).split("-")
                eff = calculate_effectiveness(
                    pd.DataFrame(aktif), df_trx, targets_df,
                    bulan=int(mo), tahun=int(y),
                )
                efektivitas_data = {
                    "volume_achievement_pct": eff["volume_achievement_pct"],
                    "peserta_aktif_pct":      eff["peserta_aktif_pct"],
                    "efektivitas_pct":        eff["efektivitas_pct"],
                    "interpretasi":           eff["interpretasi"],
                }
            except Exception:
                pass

    active_ids = {str(m["id_toko"]) for m in aktif}
    ilp_recs   = get_ilp_recommendations(crs, active_ids, limit=50)

    return {
        "status": "ok",
        "data": {
            "total_aktif":          len(aktif),
            "total_nonaktif":       len(nonaktif),
            "est_budget_bulan":     round(est_budget_total),
            "per_reward_type":      per_reward_type,
            "rekomendasi_takeout":  takeout_count,
            "rekomendasi_takein":   len(ilp_recs),
            "efektivitas_bulan_ini": efektivitas_data or None,
        },
        "meta": _meta(),
    }


# ── POST /api/loyalty/members/add-one ────────────────────────────────────────

@router.post("/members/add-one", status_code=201)
def add_one_member(body: AddOneMember) -> dict:
    if body.reward_type not in VALID_REWARD_TYPES:
        raise HTTPException(400, detail=f"reward_type tidak valid: {body.reward_type}")
    if body.cluster_pareto not in VALID_CLUSTERS:
        raise HTTPException(400, detail=f"cluster_pareto tidak valid: {body.cluster_pareto}")

    new_id = str(uuid.uuid4())
    new_member: dict = {
        "id":            new_id,
        "id_toko":       body.id_toko,
        "nama_toko":     body.nama_toko,
        "kabupaten":     body.kabupaten,
        "cluster_pareto": body.cluster_pareto,
        "tso":           body.tso,
        "reward_type":   body.reward_type,
        "catatan":       body.catatan,
        "status":        "Aktif",
        "tgl_masuk":     date.today().isoformat(),
        "tgl_keluar":    None,
        "alasan_keluar": None,
    }

    if not _add_member_if_not_duplicate(new_member):
        raise HTTPException(409, detail=f"Toko {body.id_toko} sudah ada di program loyalty")

    _log({
        "id_member": new_id, "id_toko": body.id_toko, "nama_toko": body.nama_toko,
        "tanggal": _now(), "perubahan": "Ditambahkan ke program",
        "alasan": body.catatan or "-", "status_baru": "Aktif",
    })

    return {"status": "ok", "data": new_member}


# ── POST /api/loyalty/members/upload-excel ────────────────────────────────────

@router.post("/members/upload-excel")
def upload_excel(file: UploadFile = File(...)) -> dict:
    fname = file.filename or ""
    if not fname.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, detail="File harus .xlsx atau .xls")

    contents = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
    except Exception as exc:
        raise HTTPException(400, detail=f"File tidak dapat dibaca: {exc}") from exc

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, detail="File kosong atau tidak ada baris data")

    header = [str(c or "").strip() for c in rows[0]]
    REQUIRED = {"ID Toko", "Nama Toko", "Kabupaten", "Cluster Pareto", "TSO"}
    missing  = REQUIRED - set(header)
    if missing:
        raise HTTPException(400, detail=f"Kolom tidak ditemukan: {', '.join(sorted(missing))}")

    col_idx = {name: header.index(name) for name in header}

    def gcol(row: tuple, name: str, default: str = "") -> str:
        idx = col_idx.get(name)
        return str(row[idx] or "").strip() if idx is not None and idx < len(row) else default

    berhasil   = 0
    duplikat   = 0
    errors:    list[str] = []
    new_members: list[dict] = []
    hist_events: list[dict] = []

    active_ids = _get_active_ids()

    for i, row in enumerate(rows[1:], start=2):
        id_toko = gcol(row, "ID Toko")
        nama    = gcol(row, "Nama Toko")
        kab     = gcol(row, "Kabupaten")
        cluster = gcol(row, "Cluster Pareto")
        tso     = gcol(row, "TSO")
        rt      = gcol(row, "Reward Type", "Standard")

        if not id_toko or not nama:
            errors.append(f"Baris {i}: ID Toko atau Nama Toko kosong")
            continue
        if id_toko in active_ids:
            duplikat += 1
            continue
        if cluster not in VALID_CLUSTERS:
            errors.append(f"Baris {i} ({nama}): Cluster tidak valid — '{cluster}'")
            continue
        if rt not in VALID_REWARD_TYPES:
            rt = "Standard"

        new_id = str(uuid.uuid4())
        m = {
            "id": new_id, "id_toko": id_toko, "nama_toko": nama,
            "kabupaten": kab, "cluster_pareto": cluster, "tso": tso,
            "reward_type": rt, "catatan": "",
            "status": "Aktif", "tgl_masuk": date.today().isoformat(),
            "tgl_keluar": None, "alasan_keluar": None,
        }
        new_members.append(m)
        active_ids.add(id_toko)
        berhasil += 1
        hist_events.append({
            "id_member": new_id, "id_toko": id_toko, "nama_toko": nama,
            "tanggal": _now(), "perubahan": "Upload Excel",
            "alasan": "-", "status_baru": "Aktif",
        })

    _add_members_bulk(new_members)
    _add_history_events_bulk(hist_events)

    return {
        "status": "ok",
        "data":  {"berhasil": berhasil, "duplikat": duplikat, "error": errors},
        "meta":  _meta(),
    }


# ── POST /api/loyalty/members/{id}/take-out ───────────────────────────────────

@router.post("/members/{member_id}/take-out")
def take_out_member(member_id: str, body: TakeOutBody) -> dict:
    existing = _get_member_by_id(member_id)
    if existing is None:
        raise HTTPException(404, detail="Member tidak ditemukan")
    if existing.get("status") == "Nonaktif":
        raise HTTPException(400, detail="Member sudah nonaktif")

    updated = _update_member_by_id(member_id, {
        "status":        "Nonaktif",
        "tgl_keluar":    date.today().isoformat(),
        "alasan_keluar": body.alasan,
    })

    _log({
        "id_member": member_id,
        "id_toko":   updated["id_toko"],
        "nama_toko": updated["nama_toko"],
        "tanggal":   _now(),
        "perubahan": "Take Out",
        "alasan":    body.alasan,
        "catatan":   body.catatan,
        "status_baru": "Nonaktif",
    })
    return {"status": "ok", "data": updated}


# ── PATCH /api/loyalty/members/{id}/reward-type ───────────────────────────────

@router.patch("/members/{member_id}/reward-type")
def update_reward_type(member_id: str, body: UpdateRewardTypeBody) -> dict:
    if body.reward_type not in VALID_REWARD_TYPES:
        raise HTTPException(400, detail=f"reward_type tidak valid: {body.reward_type}")

    existing = _get_member_by_id(member_id)
    if existing is None:
        raise HTTPException(404, detail="Member tidak ditemukan")
    old_rt = existing.get("reward_type", "Standard")
    updated = _update_member_by_id(member_id, {"reward_type": body.reward_type})

    _log({
        "id_member": member_id,
        "id_toko":   updated["id_toko"],
        "nama_toko": updated["nama_toko"],
        "tanggal":   _now(),
        "perubahan": f"Reward Type: {old_rt} → {body.reward_type}",
        "alasan":    "-",
        "status_baru": updated.get("status", "Aktif"),
    })
    return {"status": "ok", "data": updated}


# ── GET /api/loyalty/takeout-recommendations ──────────────────────────────────

@router.get("/takeout-recommendations")
def get_takeout_recs() -> dict:
    members = _get_members()

    active = [m for m in members if m.get("status") == "Aktif"]
    if not active:
        return {"status": "ok", "data": [], "meta": _meta(total=0)}

    crs  = get_store_crs()
    recs = get_takeout_recommendations(pd.DataFrame(active), load_data(), crs)
    data = recs.to_dict("records") if not recs.empty else []
    return {"status": "ok", "data": data, "meta": _meta(total=len(data))}


# ── GET /api/loyalty/smart-promotions ────────────────────────────────────────

@router.get("/smart-promotions")
def get_smart_promos() -> dict:
    from api.core.cannibalization_engine import load_cached_result

    members = _get_members()

    active = [m for m in members if m.get("status") == "Aktif"]
    if not active:
        return {"status": "ok", "data": [], "meta": _meta(total=0)}

    crs        = get_store_crs()
    gmm_result = load_cached_result()  # None if model not yet trained
    promos     = get_smart_promotions(pd.DataFrame(active), load_data(), crs, gmm_result=gmm_result)
    data       = promos.to_dict("records") if not promos.empty else []
    return {
        "status": "ok",
        "data":   data,
        "meta":   _meta(total=len(data), gmm_adjustment_active=gmm_result is not None),
    }


# ── GET /api/loyalty/ilp-recommendations ─────────────────────────────────────

@router.get("/ilp-recommendations")
def get_ilp_recs(limit: int = Query(50, ge=1, le=200)) -> dict:
    members = _get_members()

    active_ids = {str(m["id_toko"]) for m in members if m.get("status") == "Aktif"}
    crs  = get_store_crs()
    recs = get_ilp_recommendations(crs, active_ids, limit=limit)

    if recs.empty:
        return {"status": "ok", "data": [], "meta": _meta(total=0)}

    data: list[dict] = []
    for _, row in recs.iterrows():
        data.append({
            "id_toko":        str(row.get("ID Toko", "")),
            "nama_toko":      str(row.get("Nama Toko", "")),
            "kabupaten":      str(row.get("Kabupaten Toko", "")),
            "cluster_pareto": str(row.get("Cluster Pareto", "")),
            "tso":            str(row.get("TSO", "")),
            "ilp_score":        round(float(row.get("score", 0)), 2),
            "aegis_score":      round(float(row.get("aegis_score", 0)), 2),
            "aegis_level":      str(row.get("alert", "Normal")),
            "avg_ton_bulanan":        round(float(row.get("avg_ton", 0)), 2),
            "avg_ton_elang_bulanan":  round(float(row.get("avg_ton_elang", 0)), 2),
            "avg_ton_badak_bulanan":  round(float(row.get("avg_ton_badak", 0)), 2),
            "est_cost_bln":           round(float(row.get("estimated_cost", 0)) / 12),
        })

    return {"status": "ok", "data": data, "meta": _meta(total=len(data))}


# ── GET /api/loyalty/history ──────────────────────────────────────────────────

@router.get("/history")
def get_history() -> dict:
    hist = _get_history()
    hist_sorted = sorted(hist, key=lambda x: x.get("tanggal", ""), reverse=True)
    return {"status": "ok", "data": hist_sorted, "meta": _meta(total=len(hist_sorted))}


# ── GET /api/loyalty/template ─────────────────────────────────────────────────

@router.get("/template")
def download_template() -> StreamingResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Loyalty"

    headers = ["ID Toko", "Nama Toko", "Kabupaten", "Cluster Pareto", "TSO", "Reward Type"]
    ws.append(headers)
    ws.append(["TK001", "TOKO MAJU JAYA",    "KABUPATEN BANDUNG",  "Gold",     "TSO-01 BUDI", "Standard"])
    ws.append(["TK002", "CV SEJAHTERA",       "KOTA SURABAYA",     "Platinum", "TSO-02 SITI", "Standard"])
    ws.append(["TK003", "UD BERKAH MAKMUR",   "KABUPATEN SEMARANG", "Silver",  "TSO-03 RUDI", "Standard"])

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1D4ED8")
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill

    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col) + 4
        ws.column_dimensions[col[0].column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=loyalty_template.xlsx"},
    )


# ── GET /api/loyalty/search-stores ───────────────────────────────────────────

@router.get("/search-stores")
def search_stores(q: str = Query("", min_length=1)) -> dict:
    crs = get_store_crs()
    members = _get_members()
    active_ids = {str(m["id_toko"]) for m in members if m.get("status") == "Aktif"}

    q_lower = q.lower().strip()
    mask = (
        crs["Nama Toko"].str.lower().str.contains(q_lower, na=False)
        | crs["Kabupaten Toko"].str.lower().str.contains(q_lower, na=False)
    )
    matched = crs[mask].head(10)

    ilp     = get_ilp_features()
    ilp_idx = (
        ilp.set_index("ID Toko")
        if not ilp.empty and "ID Toko" in ilp.columns
        else pd.DataFrame()
    )

    data: list[dict] = []
    for _, row in matched.iterrows():
        id_toko = str(row["ID Toko"])
        has_ilp = not ilp_idx.empty and id_toko in ilp_idx.index
        avg_ton       = float(ilp_idx.at[id_toko, "avg_ton"]       or 0) if has_ilp else 0.0
        avg_ton_elang = float(ilp_idx.at[id_toko, "avg_ton_elang"] or 0) if has_ilp else 0.0
        avg_ton_badak = float(ilp_idx.at[id_toko, "avg_ton_badak"] or 0) if has_ilp else 0.0
        data.append({
            "id_toko":               id_toko,
            "nama_toko":             str(row.get("Nama Toko") or ""),
            "kabupaten":             str(row.get("Kabupaten Toko") or ""),
            "cluster_pareto":        str(row.get("Cluster Pareto") or ""),
            "tso":                   str(row.get("TSO") or ""),
            "aegis_score":           round(float(row.get("aegis_score") or 0), 2),
            "aegis_level":           str(row.get("alert") or "Normal"),
            "avg_ton_bulanan":        round(avg_ton, 2),
            "avg_ton_elang_bulanan":  round(avg_ton_elang, 2),
            "avg_ton_badak_bulanan":  round(avg_ton_badak, 2),
            "sudah_ada":             id_toko in active_ids,
        })

    return {"status": "ok", "data": data}


# ── GET /api/loyalty/targets ──────────────────────────────────────────────────

_EMPTY_SUMMARY = {
    "on_track": 0, "at_risk": 0, "below_target": 0,
    "avg_achievement_pct": 0.0, "total_target_ton": 0.0, "total_realisasi_ton": 0.0,
}


@router.get("/targets")
def get_targets(
    bulan:   str | None = Query(None),
    cluster: str | None = Query(None),
    status:  str | None = Query(None),
    search:  str | None = Query(None),
    limit:   int        = Query(50, ge=1, le=500),
    offset:  int        = Query(0, ge=0),
) -> dict:
    members = _get_members()
    active = [m for m in members if m.get("status") == "Aktif"]
    if not active:
        return {"status": "ok", "data": [], "summary": _EMPTY_SUMMARY,
                "meta": _meta(total=0, limit=limit, offset=offset)}

    cfg    = _load_config()
    crs    = get_store_crs()
    df_trx = load_data()

    # Single month → exact range; no month → last 12 months
    df_tmp = df_trx.copy()
    df_tmp["_p"] = df_tmp["Tanggal Transaksi"].dt.to_period("M")
    latest_p = df_tmp["_p"].max()
    if bulan:
        _start = _end = bulan
    else:
        _end   = str(latest_p)
        _start = str(latest_p - 11)

    hist = calculate_historical_targets(
        pd.DataFrame(active), df_trx, crs, config=cfg,
        bulan_start=_start, bulan_end=_end,
    )
    if not hist:
        return {"status": "ok", "data": [], "summary": _EMPTY_SUMMARY,
                "meta": _meta(total=0, limit=limit, offset=offset)}

    # Enrich with kabupaten, tso, aegis_level, aegis_score from members/CRS
    member_map = {str(m["id_toko"]): m for m in active}
    crs_idx    = crs.set_index("ID Toko") if "ID Toko" in crs.columns else pd.DataFrame()
    for row in hist:
        id_t = str(row["id_toko"])
        mem  = member_map.get(id_t, {})
        row["kabupaten"] = str(mem.get("kabupaten", ""))
        row["tso"]       = str(mem.get("tso", ""))
        if not crs_idx.empty and id_t in crs_idx.index:
            row["aegis_level"] = str(crs_idx.at[id_t, "alert"]       or "Normal")
            row["aegis_score"] = round(float(crs_idx.at[id_t, "aegis_score"] or 0), 2)
        else:
            row["aegis_level"] = "Normal"
            row["aegis_score"] = 0.0

    # Apply filters
    if cluster:
        hist = [r for r in hist if r["cluster"] == cluster]
    if status:
        hist = [r for r in hist if r["status_achievement"] == status]
    if search:
        sq = search.lower()
        hist = [r for r in hist if sq in r["id_toko"].lower() or sq in r["nama_toko"].lower()]

    total = len(hist)
    on_track      = sum(1 for r in hist if r["status_achievement"] == "On Track")
    at_risk       = sum(1 for r in hist if r["status_achievement"] == "At Risk")
    below_target  = sum(1 for r in hist if r["status_achievement"] == "Below Target")
    avg_ach       = (sum(r["achievement_pct"] for r in hist) / total) if total else 0.0
    total_tgt     = sum(r["target_ton"]    for r in hist)
    total_real    = sum(r["realisasi_ton"] for r in hist)

    summary = {
        "on_track":             on_track,
        "at_risk":              at_risk,
        "below_target":         below_target,
        "avg_achievement_pct":  round(avg_ach, 2),
        "total_target_ton":     round(total_tgt, 2),
        "total_realisasi_ton":  round(total_real, 2),
    }

    page = hist[offset: offset + limit]
    return {"status": "ok", "data": page, "summary": summary,
            "meta": _meta(total=total, limit=limit, offset=offset)}


# ── GET /api/loyalty/targets/export ───────────────────────────────────────────

@router.get("/targets/export")
def export_targets(
    bulan:   str | None = Query(None),
    cluster: str | None = Query(None),
    status:  str | None = Query(None),
    search:  str | None = Query(None),
) -> StreamingResponse:
    """Export filtered historical targets as .xlsx."""
    members = _get_members()
    active = [m for m in members if m.get("status") == "Aktif"]
    if not active:
        raise HTTPException(404, detail="Tidak ada peserta aktif")

    cfg    = _load_config()
    crs    = get_store_crs()
    df_trx = load_data()

    df_tmp = df_trx.copy()
    df_tmp["_p"] = df_tmp["Tanggal Transaksi"].dt.to_period("M")
    latest_p = df_tmp["_p"].max()
    if bulan:
        _start = _end = bulan
    else:
        _end   = str(latest_p)
        _start = str(latest_p - 11)

    hist = calculate_historical_targets(
        pd.DataFrame(active), df_trx, crs, config=cfg,
        bulan_start=_start, bulan_end=_end,
    )

    # Enrich
    member_map = {str(m["id_toko"]): m for m in active}
    crs_idx    = crs.set_index("ID Toko") if "ID Toko" in crs.columns else pd.DataFrame()
    for row in hist:
        id_t = str(row["id_toko"])
        mem  = member_map.get(id_t, {})
        row["kabupaten"] = str(mem.get("kabupaten", ""))
        row["tso"]       = str(mem.get("tso", ""))
        if not crs_idx.empty and id_t in crs_idx.index:
            row["aegis_level"] = str(crs_idx.at[id_t, "alert"]       or "Normal")
        else:
            row["aegis_level"] = "Normal"

    # Apply filters
    if cluster:
        hist = [r for r in hist if r["cluster"] == cluster]
    if status:
        hist = [r for r in hist if r["status_achievement"] == status]
    if search:
        sq = search.lower()
        hist = [r for r in hist if sq in r["id_toko"].lower() or sq in r["nama_toko"].lower()]

    # Build Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Target Achievement"

    headers = ["Periode", "ID Toko", "Nama Toko", "AEGIS", "Cluster", "Kabupaten", "TSO",
               "Target TON", "Realisasi TON", "ACH%", "Status"]
    ws.append(headers)

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1D4ED8")
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill

    STATUS_COLORS = {
        "On Track":     "16A34A",
        "At Risk":      "D97706",
        "Below Target": "DC2626",
    }
    for row in hist:
        ws.append([
            row.get("periode_label", row.get("periode", "")),
            row["id_toko"],
            row["nama_toko"],
            row.get("aegis_level", "Normal"),
            row["cluster"],
            row.get("kabupaten", ""),
            row.get("tso", ""),
            round(row["target_ton"], 2),
            round(row["realisasi_ton"], 2),
            round(row["achievement_pct"], 2),
            row["status_achievement"],
        ])
        # Color-code status cell
        status_val = row["status_achievement"]
        color      = STATUS_COLORS.get(status_val)
        if color:
            ws.cell(row=ws.max_row, column=11).font = Font(color=color, bold=True)

    for col in ws.columns:
        width = max(len(str(cell.value or "")) for cell in col) + 3
        ws.column_dimensions[col[0].column_letter].width = min(width, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"targets_{bulan or f'{_start}_to_{_end}'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ── GET /api/loyalty/targets/summary ─────────────────────────────────────────

@router.get("/targets/summary")
def get_targets_summary() -> dict:
    members = _get_members()
    active = [m for m in members if m.get("status") == "Aktif"]
    if not active:
        return {
            "status": "ok",
            "data": {
                "total": 0, "avg_achievement_pct": 0.0,
                "on_track": 0, "at_risk": 0, "below_target": 0,
                "lowest_achievers": [], "bulan_target": None,
            },
            "meta": _meta(),
        }

    cfg    = _load_config()
    crs    = get_store_crs()
    df_trx = load_data()
    targets_df = calculate_loyalty_targets(pd.DataFrame(active), df_trx, crs, config=cfg)

    if targets_df.empty:
        return {
            "status": "ok",
            "data": {
                "total": 0, "avg_achievement_pct": 0.0,
                "on_track": 0, "at_risk": 0, "below_target": 0,
                "lowest_achievers": [], "bulan_target": None,
            },
            "meta": _meta(),
        }

    triggers_df  = get_target_triggers(targets_df)
    lowest_five  = (
        targets_df.nsmallest(5, "achievement_pct")[
            ["id_toko", "nama_toko", "cluster_pareto", "achievement_pct", "status"]
        ]
        .where(pd.notna(targets_df.nsmallest(5, "achievement_pct")), None)
        .to_dict("records")
    )

    return {
        "status": "ok",
        "data": {
            "total":               len(targets_df),
            "avg_achievement_pct": round(float(targets_df["achievement_pct"].mean()), 2),
            "on_track":            int((targets_df["status"] == "On Track").sum()),
            "at_risk":             int((targets_df["status"] == "At Risk").sum()),
            "below_target":        int((targets_df["status"] == "Below Target").sum()),
            "triggers":            len(triggers_df),
            "lowest_achievers":    lowest_five,
            "bulan_target":        str(targets_df["bulan_target"].iloc[0]),
        },
        "meta": _meta(),
    }


# ── GET /api/loyalty/targets/config ──────────────────────────────────────────

@router.get("/targets/config")
def get_targets_config() -> dict:
    return {"status": "ok", "data": _load_config(), "meta": _meta()}


# ── POST /api/loyalty/targets/config ─────────────────────────────────────────

@router.post("/targets/config")
def update_targets_config(body: TargetConfigBody) -> dict:
    if abs(body.w1 + body.w2 - 1.0) > 0.01:
        raise HTTPException(400, detail="w1 + w2 harus = 1.0")

    d = body.growth_rates.default
    for field, val in [("normal", d.normal), ("warning", d.warning), ("kritis", d.kritis)]:
        if not (0.0 <= val <= 0.50):
            raise HTTPException(400, detail=f"growth_rates.default.{field} harus antara 0% dan 50%")

    for i, ov in enumerate(body.growth_rates.overrides):
        if ov.type == "monthly":
            if ov.bulan is None or not (1 <= ov.bulan <= 12):
                raise HTTPException(400, detail=f"Override {i}: bulan harus 1–12")
        elif ov.type == "quarterly":
            if ov.kuartal is None or not (1 <= ov.kuartal <= 4):
                raise HTTPException(400, detail=f"Override {i}: kuartal harus 1–4")
        else:
            raise HTTPException(400, detail=f"Override {i}: type harus 'monthly' atau 'quarterly'")
        if ov.tahun < 2024:
            raise HTTPException(400, detail=f"Override {i}: tahun minimal 2024")
        for field, val in [("normal", ov.normal), ("warning", ov.warning), ("kritis", ov.kritis)]:
            if not (0.0 <= val <= 0.50):
                raise HTTPException(400, detail=f"Override {i}: {field} harus antara 0% dan 50%")

    for field, val in [
        ("min_pct_sp", body.min_pct_sp), ("min_pct_platinum", body.min_pct_platinum),
        ("min_pct_gold", body.min_pct_gold), ("min_pct_silver", body.min_pct_silver),
    ]:
        if not (0.0 <= val <= 1.0):
            raise HTTPException(400, detail=f"{field} harus antara 0 dan 1")

    new_cfg = body.model_dump()
    _save_target_config(new_cfg)
    return {"status": "ok", "data": new_cfg, "meta": _meta()}


# ── GET /api/loyalty/targets/growth-schedule ──────────────────────────────────

@router.get("/targets/growth-schedule")
def get_growth_schedule() -> dict:
    """Return growth-rate calendar for the next 12 months from the latest data month."""
    cfg              = _load_config()
    growth_rates_cfg = cfg.get("growth_rates", DEFAULT_CONFIG["growth_rates"])

    df = load_data()
    df["_p"] = df["Tanggal Transaksi"].dt.to_period("M")
    start_p = df["_p"].max()

    month_names = ["Jan","Feb","Mar","Apr","Mei","Jun","Jul","Agu","Sep","Okt","Nov","Des"]

    schedule: list[dict] = []
    for i in range(12):
        p     = start_p + i
        b, y  = p.month, p.year
        n_rate, sumber = get_growth_rate("normal",  b, y, growth_rates_cfg)
        w_rate, _      = get_growth_rate("warning", b, y, growth_rates_cfg)
        k_rate, _      = get_growth_rate("kritis",  b, y, growth_rates_cfg)
        schedule.append({
            "bulan":       f"{month_names[b - 1]} {y}",
            "bulan_num":   b,
            "tahun":       y,
            "normal_pct":  round(n_rate * 100, 2),
            "warning_pct": round(w_rate * 100, 2),
            "kritis_pct":  round(k_rate * 100, 2),
            "sumber":      sumber,
        })

    return {"status": "ok", "data": schedule, "meta": _meta()}


# ── GET /api/loyalty/targets/history ──────────────────────────────────────────

@router.get("/targets/history")
def get_targets_history(
    bulan_start: str = Query("2024-01"),
    bulan_end:   str = Query("2026-04"),
) -> dict:
    """Historical targets for all active members from bulan_start to bulan_end."""
    members = _get_members()
    active = [m for m in members if m.get("status") == "Aktif"]
    if not active:
        return {"status": "ok", "data": [], "meta": _meta(total=0)}

    cfg    = _load_config()
    crs    = get_store_crs()
    df_trx = load_data()

    hist = calculate_historical_targets(
        pd.DataFrame(active), df_trx, crs, config=cfg,
        bulan_start=bulan_start, bulan_end=bulan_end,
    )
    if not hist:
        return {"status": "ok", "data": [], "meta": _meta(total=0)}

    # Group by periode
    from collections import defaultdict
    by_period: dict[str, list[dict]] = defaultdict(list)
    for row in hist:
        by_period[row["periode"]].append(row)

    result: list[dict] = []
    for periode in sorted(by_period.keys()):
        rows  = by_period[periode]
        label = rows[0]["periode_label"]
        total_target   = sum(r["target_ton"]  for r in rows)
        total_realisasi = sum(r["realisasi_ton"] for r in rows)
        avg_ach = (
            sum(r["achievement_pct"] for r in rows) / len(rows)
            if rows else 0.0
        )
        on_track   = sum(1 for r in rows if r["status_achievement"] == "On Track")
        at_risk    = sum(1 for r in rows if r["status_achievement"] == "At Risk")
        below      = sum(1 for r in rows if r["status_achievement"] == "Below Target")
        result.append({
            "periode":            periode,
            "label":              label,
            "total_target":       round(total_target, 2),
            "total_realisasi":    round(total_realisasi, 2),
            "avg_achievement_pct": round(avg_ach, 2),
            "on_track_count":     on_track,
            "at_risk_count":      at_risk,
            "below_target_count": below,
            "stores":             rows,
        })

    return {"status": "ok", "data": result, "meta": _meta(total=len(result))}


# ── GET /api/loyalty/effectiveness ────────────────────────────────────────────

@router.get("/effectiveness")
def get_effectiveness_endpoint(
    bulan: int = Query(..., ge=1, le=12),
    tahun: int = Query(..., ge=2020),
) -> dict:
    members = _get_members()
    active = [m for m in members if m.get("status") == "Aktif"]
    if not active:
        return {"status": "ok", "data": calculate_effectiveness(pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), bulan, tahun), "meta": _meta()}

    cfg    = _load_config()
    crs    = get_store_crs()
    df_trx = load_data()

    df_tmp = df_trx.copy()
    df_tmp["_p"] = df_tmp["Tanggal Transaksi"].dt.to_period("M")
    latest_p = df_tmp["_p"].max()

    if pd.Period(f"{tahun}-{bulan:02d}", "M") == latest_p:
        targets_df = calculate_loyalty_targets(pd.DataFrame(active), df_trx, crs, config=cfg)
    else:
        targets_df = pd.DataFrame()

    eff = calculate_effectiveness(pd.DataFrame(active), df_trx, targets_df, bulan, tahun)
    return {"status": "ok", "data": eff, "meta": _meta()}


# ── GET /api/loyalty/insights/volume-trend ────────────────────────────────────

@router.get("/insights/volume-trend")
def get_insights_volume_trend() -> dict:
    """12-month volume trend: loyalty vs non-loyalty, with targets and achievement %."""
    from collections import defaultdict

    members = _get_members()
    active = [m for m in members if m.get("status") == "Aktif"]
    active_ids = {str(m["id_toko"]) for m in active}

    cfg    = _load_config()
    crs    = get_store_crs()
    df_trx = load_data()

    df = df_trx.copy()
    df["_p"] = df["Tanggal Transaksi"].dt.to_period("M")
    latest_p = df["_p"].max()
    start_p  = latest_p - 11

    agg_dict: dict = df.groupby(["ID Toko", "_p"])["TON Quantity"].sum().to_dict()

    # Stores by period index for non-loyalty lookup
    stores_by_period: dict = defaultdict(set)
    for (store, period) in agg_dict.keys():
        stores_by_period[period].add(store)

    # Historical targets aggregated by period
    target_by_period: dict[str, float] = defaultdict(float)
    if active:
        hist = calculate_historical_targets(
            pd.DataFrame(active), df_trx, crs, config=cfg,
            bulan_start=str(start_p), bulan_end=str(latest_p),
        )
        for row in hist:
            target_by_period[row["periode"]] += row["target_ton"]

    month_names = ["Jan","Feb","Mar","Apr","Mei","Jun","Jul","Agu","Sep","Okt","Nov","Des"]
    result: list[dict] = []
    for i in range(12):
        p = start_p + i
        b, y = p.month, p.year

        vol_loyalty = sum(agg_dict.get((sid, p), 0.0) for sid in active_ids)
        non_loyalty_ids = stores_by_period.get(p, set()) - active_ids
        vol_non_loyalty = sum(agg_dict.get((sid, p), 0.0) for sid in non_loyalty_ids)
        tgt     = target_by_period.get(str(p), 0.0)
        ach_pct = (vol_loyalty / tgt * 100) if tgt > 0 else 0.0

        result.append({
            "bulan":             str(p),
            "bulan_label":       f"{month_names[b-1]} {y}",
            "volume_loyalty":    round(vol_loyalty, 2),
            "volume_non_loyalty": round(vol_non_loyalty, 2),
            "target_loyalty":    round(tgt, 2),
            "achievement_pct":   round(ach_pct, 2),
        })

    return {"status": "ok", "data": result, "meta": _meta()}


# ── GET /api/loyalty/insights/comparison ──────────────────────────────────────

@router.get("/insights/comparison")
def get_insights_comparison() -> dict:
    """6-month avg TON per store: loyalty vs non-loyalty."""
    from collections import defaultdict

    members = _get_members()
    active = [m for m in members if m.get("status") == "Aktif"]
    active_ids = {str(m["id_toko"]) for m in active}

    df_trx = load_data()
    df = df_trx.copy()
    df["_p"] = df["Tanggal Transaksi"].dt.to_period("M")
    latest_p = df["_p"].max()
    start_p  = latest_p - 5

    agg_dict: dict = df.groupby(["ID Toko", "_p"])["TON Quantity"].sum().to_dict()
    stores_by_period: dict = defaultdict(set)
    for (store, period) in agg_dict.keys():
        stores_by_period[period].add(store)

    month_names = ["Jan","Feb","Mar","Apr","Mei","Jun","Jul","Agu","Sep","Okt","Nov","Des"]
    result: list[dict] = []
    ratios: list[float] = []

    for i in range(6):
        p = start_p + i
        b, y = p.month, p.year

        loyalty_vols = [agg_dict.get((sid, p), 0.0) for sid in active_ids if agg_dict.get((sid, p), 0.0) > 0]
        non_ids      = stores_by_period.get(p, set()) - active_ids
        non_vols     = [agg_dict.get((sid, p), 0.0) for sid in non_ids if agg_dict.get((sid, p), 0.0) > 0]

        avg_loyalty     = float(sum(loyalty_vols) / len(loyalty_vols)) if loyalty_vols else 0.0
        avg_non_loyalty = float(sum(non_vols)     / len(non_vols))     if non_vols     else 0.0
        ratio           = (avg_loyalty / avg_non_loyalty) if avg_non_loyalty > 0 else 0.0
        if ratio > 0:
            ratios.append(ratio)

        result.append({
            "bulan":             str(p),
            "bulan_label":       f"{month_names[b-1]} {y}",
            "avg_ton_loyalty":   round(avg_loyalty, 2),
            "avg_ton_non_loyalty": round(avg_non_loyalty, 2),
            "ratio":             round(ratio, 2),
            "loyalty_toko_count":   len(loyalty_vols),
            "non_loyalty_toko_count": len(non_vols),
        })

    avg_ratio = round(float(sum(ratios) / len(ratios)), 2) if ratios else 0.0
    return {"status": "ok", "data": result, "meta": _meta(avg_ratio=avg_ratio)}


# ── GET /api/loyalty/insights/effectiveness-trend ─────────────────────────────

@router.get("/insights/effectiveness-trend")
def get_insights_effectiveness_trend() -> dict:
    """12-month effectiveness trend."""
    from collections import defaultdict

    members = _get_members()
    active = [m for m in members if m.get("status") == "Aktif"]
    active_ids = {str(m["id_toko"]) for m in active}
    total_peserta = len(active)

    cfg    = _load_config()
    crs    = get_store_crs()
    df_trx = load_data()

    df = df_trx.copy()
    df["_p"] = df["Tanggal Transaksi"].dt.to_period("M")
    latest_p = df["_p"].max()
    start_p  = latest_p - 11

    agg_dict: dict = df.groupby(["ID Toko", "_p"])["TON Quantity"].sum().to_dict()

    # Historical targets for volume achievement
    target_by_period: dict[str, float] = defaultdict(float)
    realisasi_by_period: dict[str, float] = defaultdict(float)
    if active:
        hist = calculate_historical_targets(
            pd.DataFrame(active), df_trx, crs, config=cfg,
            bulan_start=str(start_p), bulan_end=str(latest_p),
        )
        for row in hist:
            target_by_period[row["periode"]]    += row["target_ton"]
            realisasi_by_period[row["periode"]] += row["realisasi_ton"]

    month_names = ["Jan","Feb","Mar","Apr","Mei","Jun","Jul","Agu","Sep","Okt","Nov","Des"]
    result: list[dict] = []

    for i in range(12):
        p = start_p + i
        b, y = p.month, p.year

        tgt  = target_by_period.get(str(p), 0.0)
        real = realisasi_by_period.get(str(p), 0.0)
        if real == 0.0:
            real = sum(agg_dict.get((sid, p), 0.0) for sid in active_ids)
        volume_ach = (real / tgt * 100) if tgt > 0 else 0.0

        peserta_bertransaksi = sum(1 for sid in active_ids if agg_dict.get((sid, p), 0.0) > 0)
        aktif_pct = (peserta_bertransaksi / total_peserta * 100) if total_peserta > 0 else 0.0

        efektivitas = (volume_ach * 0.6) + (aktif_pct * 0.4)

        result.append({
            "bulan":                  str(p),
            "bulan_label":            f"{month_names[b-1]} {y}",
            "volume_achievement_pct": round(volume_ach, 2),
            "peserta_aktif_pct":      round(aktif_pct, 2),
            "efektivitas_pct":        round(efektivitas, 2),
        })

    return {"status": "ok", "data": result, "meta": _meta()}
